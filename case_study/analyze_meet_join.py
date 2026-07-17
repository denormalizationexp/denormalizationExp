#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
analyze_meet_join.py
====================
Re-slice the unified per-query results into the *meet / join* view-selection
experiment for a set of small workloads W (Sec. 4.3 "Understanding View
Selection Conceptually").  For a workload W (a set of join queries Q_j), the
lattice supplies two canonical single-view denormalizations:

    meet  $\\sqcap W$  = the *largest* view every query in W contains
                       (atoms = intersection of the queries' atoms); each
                       query becomes a reduced join (Q - V) join V.
    join  $\\sqcup W$  = the *smallest* view that contains every query in W
                       (atoms = union of the queries' atoms); each query is a
                       projection of V.

Because Q_j is the join defining lattice node j, atoms(Q_j) = atoms(node j),
so meet(W) and join(W) are simply the bit-AND / bit-OR of the query indices'
atom sets, i.e. two lattice nodes DB_m / DB_j whose per-query times we already
measured for Table~\\ref{tab:results}.  No new database run is needed: the
comparison is a re-slice of the SAME timings behind the main results, so the
numbers are mutually consistent with the rest of Sec. 6.

For each (workload, view) we report, per scale factor:
    speedup  = wl_W(DB0) / wl_W(view)          (>1 faster, <1 slower)
    kappa_W  = ceil( create(view) / (wl_W(DB0) - wl_W(view)) )   [inf if <=0]
and the (workload-independent) redundancy of the view vs DB0.

Usage:
    python analyze_meet_join.py --xlsx denorm_full_results_with_fk.xlsx
    python analyze_meet_join.py --xlsx denorm_full_results_with_fk.xlsx --latex
    python analyze_meet_join.py --heatmap

The heatmap mode renders the complete 22 x 15 per-query matrix: the 15
single-view strategies first, followed by the 7 complementary multi-view
strategies.  By default each cell is DB0_time - strategy_time in seconds, so a
positive (blue) value means the strategy is faster than DB0.
"""
import argparse
import math
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR / "denorm_full_results_with_fk.xlsx"
DEFAULT_FIGURE_DIR = (
    SCRIPT_DIR.parent
    / "case_study"
    / "figures"
    / "sports_case_study"
)

# atom set of each query Q_j = atoms of lattice node j (NODE_ORDER in module R)
ATOMS_OF = {
    1: {"PS"}, 2: {"Gh"}, 3: {"Ga"}, 4: {"SG"},
    5: {"PS", "Gh"}, 6: {"PS", "Ga"}, 7: {"Gh", "Ga"}, 8: {"PS", "SG"},
    9: {"SG", "Gh"}, 10: {"SG", "Ga"}, 11: {"PS", "Gh", "Ga"},
    12: {"PS", "SG", "Gh"}, 13: {"PS", "SG", "Ga"}, 14: {"SG", "Gh", "Ga"},
    15: {"PS", "SG", "Gh", "Ga"},
}
# node index <-> atom set (index i has the same atoms as Q_i; DB0 = empty)
NODE_ATOMS = {0: frozenset()}
NODE_ATOMS.update({j: frozenset(a) for j, a in ATOMS_OF.items()})
ATOMS_NODE = {v: k for k, v in NODE_ATOMS.items()}

# the workloads to study (Sec. 4.3 gives W={Q5,Q8}; the study varies the spread)
WORKLOADS = [
    ("W_1", [5, 7, 9]),
    ("W_2", [8, 9]),
    ("W_3", [12, 13]),
]

SINGLE_STRATEGIES = [str(i) for i in range(1, 16)]
MULTI_STRATEGIES = [f"{i}+{15 - i}" for i in range(7, 0, -1)]
QUERY_LABELS = [f"Q{i}" for i in range(1, 16)]


def meet_node(W):
    atoms = frozenset.intersection(*[NODE_ATOMS[j] for j in W])
    return ATOMS_NODE[atoms]


def join_node(W):
    atoms = frozenset.union(*[NODE_ATOMS[j] for j in W])
    return ATOMS_NODE[atoms]


def load(path):
    wb = load_workbook(path, data_only=True)

    def as_map(name):
        ws = wb[name]
        hdr = [c.value for c in ws[1]]
        ix: dict[str, int] = {
            n: k
            for k, n in enumerate(hdr)
            if isinstance(n, str)
        }
        D = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[ix["sf"]] is None:
                continue
            lab = str(r[ix["i"]]).replace("DB", "")
            D[(str(r[ix["sf"]]), lab)] = r
        return D, ix

    return as_map("per_query_time") + as_map("summary")


def _sf_key(value):
    """Return a stable printable key for integer or fractional scale factors."""
    value = float(value)
    return str(int(value)) if value.is_integer() else f"{value:g}"


def _pretty_node(node):
    """Turn workbook node labels into compact labels such as PS·SG | Gh·Ga."""
    return " | ".join(
        part.strip().replace(" ", "·") for part in str(node).split("|")
    )


def load_per_query_delta(path, sf, direction="saving"):
    """Load the requested 22 x 15 matrix directly from per_query_time.

    ``saving`` means DB0 - strategy (positive is faster); ``overhead`` is the
    inverse strategy - DB0 convention.  Reading the raw timings rather than the
    precomputed saving sheet keeps the direction explicit and auditable.
    """
    if direction not in {"saving", "overhead"}:
        raise ValueError("direction must be 'saving' or 'overhead'")

    wb = load_workbook(path, data_only=True, read_only=True)
    if "per_query_time" not in wb.sheetnames:
        raise ValueError(f"{path} has no 'per_query_time' sheet")

    ws = wb["per_query_time"]
    headers = [cell.value for cell in ws[1]]
    ix: dict[str, int] = {
        name: idx
        for idx, name in enumerate(headers)
        if isinstance(name, str)
    }
    required = {"sf", "i", "node", "view_type", "kind", *QUERY_LABELS}
    missing_columns = sorted(required - ix.keys())
    if missing_columns:
        raise ValueError(
            "per_query_time is missing columns: " + ", ".join(missing_columns)
        )

    records = {}
    wanted_sf = float(sf)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_sf = row[ix["sf"]]
        if row_sf is None or not math.isclose(float(row_sf), wanted_sf):
            continue
        # Pair rows are materialized-only, so use the matching materialized DB0
        # and single rows even when the workbook also contains plain views.
        if str(row[ix["view_type"]]).lower() != "materialized":
            continue
        label = str(row[ix["i"]]).replace("DB", "").strip()
        if label in records:
            raise ValueError(
                f"duplicate materialized strategy {label!r} for sf={_sf_key(sf)}"
            )
        records[label] = {
            "node": row[ix["node"]],
            "kind": row[ix["kind"]],
            "times": [float(row[ix[q]]) for q in QUERY_LABELS],
        }
    wb.close()

    if "0" in records and "DB0" not in records:
        records["DB0"] = records.pop("0")
    if "DB0" not in records:
        raise ValueError(f"no materialized DB0 row found for sf={_sf_key(sf)}")

    strategy_order = SINGLE_STRATEGIES + MULTI_STRATEGIES
    missing = [label for label in strategy_order if label not in records]
    if missing:
        raise ValueError(
            f"sf={_sf_key(sf)} is missing strategies: {', '.join(missing)}"
        )
    wrong_kinds = [
        label
        for label in SINGLE_STRATEGIES
        if records[label]["kind"] != "single"
    ] + [
        label
        for label in MULTI_STRATEGIES
        if records[label]["kind"] != "pair"
    ]
    if wrong_kinds:
        raise ValueError(
            "unexpected strategy kind for: " + ", ".join(wrong_kinds)
        )

    baseline = records["DB0"]["times"]
    sign = 1.0 if direction == "saving" else -1.0
    matrix = [
        [
            sign * (base - value)
            for base, value in zip(baseline, records[label]["times"])
        ]
        for label in strategy_order
    ]
    row_labels = [
        f"DB{label} — {_pretty_node(records[label]['node'])}"
        for label in strategy_order
    ]
    # Q_j is the join defining DB_j, so the single-view node label is also the
    # query's atom annotation.
    column_labels = [
        f"Q{j}\n{_pretty_node(records[str(j)]['node'])}" for j in range(1, 16)
    ]
    return matrix, row_labels, column_labels


def _cell_text(value):
    """Compact annotation that still shows every cell, including exact zero."""
    if value == 0:
        return "0"
    if abs(value) < 0.0001:
        return f"{value:.1e}"
    if abs(value) < 0.01:
        return f"{value:.4f}"
    if abs(value) < 0.1:
        return f"{value:.3f}"
    if abs(value) < 10:
        return f"{value:.2f}"
    return f"{value:.1f}"


def plot_per_query_delta(path, sf, output, direction="saving", dpi=200):
    """Render a numbered 22 x 15 diverging heatmap."""
    try:
        import matplotlib.pyplot as plt
        import numpy as np
        from matplotlib.colors import SymLogNorm
        from matplotlib.ticker import FuncFormatter
    except ImportError as exc:
        raise SystemExit("matplotlib and numpy are required for --heatmap") from exc

    matrix, row_labels, column_labels = load_per_query_delta(path, sf, direction)
    values = np.asarray(matrix, dtype=float)
    max_abs = float(np.max(np.abs(values)))
    if max_abs == 0:
        max_abs = 1.0
    linthresh = max(max_abs / 100.0, 1e-6)
    norm = SymLogNorm(
        linthresh=linthresh,
        linscale=1.0,
        vmin=-max_abs,
        vmax=max_abs,
        base=10,
    )
    # Colour always carries the same semantics: blue=faster, red=slower.
    cmap = plt.get_cmap("RdBu" if direction == "saving" else "RdBu_r")

    fig, ax = plt.subplots(figsize=(19.2, 11.4), constrained_layout=False)
    image = ax.imshow(values, cmap=cmap, norm=norm, aspect="auto")

    ax.set_xticks(np.arange(len(column_labels)), labels=column_labels)
    ax.set_yticks(np.arange(len(row_labels)), labels=row_labels)
    ax.tick_params(axis="x", length=0, pad=8, labelsize=9)
    ax.tick_params(axis="y", length=0, pad=7, labelsize=9)
    ax.set_xlabel(
        "Query (atoms touched)", fontsize=12, fontweight="bold", labelpad=14
    )
    ax.set_ylabel("Strategy", fontsize=12, fontweight="bold", labelpad=76)

    # Fine white rules make the output read as a table without heavy borders.
    ax.set_xticks(np.arange(-0.5, values.shape[1], 1), minor=True)
    ax.set_yticks(np.arange(-0.5, values.shape[0], 1), minor=True)
    ax.grid(which="minor", color="white", linewidth=0.8)
    ax.tick_params(which="minor", bottom=False, left=False)
    ax.axhline(len(SINGLE_STRATEGIES) - 0.5, color="#202020", linewidth=2.0)

    for row in range(values.shape[0]):
        for col in range(values.shape[1]):
            value = values[row, col]
            red, green, blue, _alpha = cmap(norm(value))
            luminance = 0.2126 * red + 0.7152 * green + 0.0722 * blue
            ax.text(
                col,
                row,
                _cell_text(value),
                ha="center",
                va="center",
                fontsize=7.4,
                fontweight="semibold",
                color="white" if luminance < 0.48 else "black",
            )

    # The row order itself carries the requested 15 Single / 7 Multi grouping.
    row_count = len(SINGLE_STRATEGIES) + len(MULTI_STRATEGIES)
    single_y = 1 - (len(SINGLE_STRATEGIES) / 2) / row_count
    multi_y = (len(MULTI_STRATEGIES) / 2) / row_count
    ax.text(
        -0.245,
        single_y,
        "Single (15)",
        transform=ax.transAxes,
        rotation=90,
        ha="center",
        va="center",
        fontsize=10,
        fontweight="bold",
    )
    ax.text(
        -0.245,
        multi_y,
        "Multi (7)",
        transform=ax.transAxes,
        rotation=90,
        ha="center",
        va="center",
        fontsize=10,
        fontweight="bold",
    )

    sf_label = _sf_key(sf)
    if direction == "saving":
        title = f"Per-query workload saving vs DB0 (sf={sf_label})"
        subtitle = "cell = DB0 time − strategy time (s); positive/blue is faster"
        colorbar_label = "DB0 time − strategy time (s)"
    else:
        title = f"Per-query workload overhead vs DB0 (sf={sf_label})"
        subtitle = "cell = strategy time − DB0 time (s); positive/red is slower"
        colorbar_label = "Strategy time − DB0 time (s)"
    fig.suptitle(title, fontsize=15, fontweight="bold", y=0.985)
    ax.set_title(subtitle, fontsize=10, pad=13)

    colorbar = fig.colorbar(image, ax=ax, fraction=0.025, pad=0.012)
    colorbar.set_label(colorbar_label, fontsize=10, fontweight="bold", labelpad=10)
    colorbar.ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _pos: f"{x:g}"))
    colorbar.ax.tick_params(labelsize=8)

    fig.subplots_adjust(left=0.245, right=0.93, top=0.91, bottom=0.12)
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output, dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return output


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--sfs", default="10,100")
    ap.add_argument("--latex", action="store_true",
                    help="emit the LaTeX table body instead of a console report")
    ap.add_argument("--heatmap", action="store_true",
                    help="render the 22 x 15 per-query delta table instead")
    ap.add_argument("--figure-dir", type=Path, default=DEFAULT_FIGURE_DIR,
                    help="directory for heatmap PNG files")
    ap.add_argument("--delta-direction", choices=["saving", "overhead"],
                    default="saving",
                    help="saving: DB0-strategy; overhead: strategy-DB0")
    ap.add_argument("--dpi", type=int, default=200,
                    help="heatmap resolution (default: 200)")
    args = ap.parse_args()

    sfs = [s.strip() for s in args.sfs.split(",") if s.strip()]
    if args.heatmap:
        for sf in sfs:
            stem = (
                "per_query_saving"
                if args.delta_direction == "saving"
                else "per_query_overhead"
            )
            output = args.figure_dir / f"{stem}_heatmap_sf{_sf_key(sf)}.png"
            plot_per_query_delta(
                args.xlsx, sf, output, args.delta_direction, dpi=args.dpi
            )
            print(f"[ok] wrote {output}")
        return

    PT, pix, SM, six = load(args.xlsx)

    def wl(sf, node, W):
        r = PT[(sf, str(node))]
        return sum(r[pix[f"Q{j}"]] for j in W)

    def create(sf, node):
        return SM[(sf, str(node))][six["create_time_s"]]

    def redund(sf, node):
        return SM[(sf, str(node))][six["redundancy_vs_db0"]]

    def kappa(sf, node, W, base):
        d = base - wl(sf, node, W)
        return math.ceil(create(sf, node) / d) if d > 1e-9 else math.inf

    def sp(sf, node, W, base):
        w = wl(sf, node, W)
        return base / w if w else float("nan")

    rows = []  # (wl_name, role, node, red, {sf: (speedup, kappa)})
    for name, W in WORKLOADS:
        m, j = meet_node(W), join_node(W)
        base = {sf: wl(sf, 0, W) for sf in sfs}
        rows.append((name, W, "normalized", 0,
                     {sf: (1.0, None) for sf in sfs}, 1.0))
        rows.append((name, W, "meet", m,
                     {sf: (sp(sf, m, W, base[sf]), kappa(sf, m, W, base[sf]))
                      for sf in sfs}, redund(sfs[0], m)))
        rows.append((name, W, "join", j,
                     {sf: (sp(sf, j, W, base[sf]), kappa(sf, j, W, base[sf]))
                      for sf in sfs}, redund(sfs[0], j)))

    if not args.latex:
        for name, W, role, node, per, red in rows:
            cells = "  ".join(
                f"sf{sf}: {per[sf][0]:5.2f}x k="
                f"{'inf' if per[sf][1] in (None, math.inf) else per[sf][1]}"
                for sf in sfs)
            wl_ms = "  ".join(f"{wl(sf, node, W)*1000:8.1f}ms" for sf in sfs)
            print(f"{name} {role:<10} DB{node:<2} red={red:4.2f}  {cells}  |  "
                  f"wl {wl_ms}")
        return

    # ---- LaTeX table body ----
    def fk(k):
        if k is None:
            return "--"
        if k is math.inf:
            return r"$\infty$"
        return str(k)

    def esc(name):
        return name.replace("W_", "W_")  # already math-ready as W_1 etc.

    print(r"% --- auto-generated by analyze_meet_join.py; do not edit by hand ---")
    for gi, (name, W) in enumerate(WORKLOADS):
        m, j = meet_node(W), join_node(W)
        qset = ",".join(f"Q_{{{q}}}" for q in W)
        hdr = (f"\\rowcolor{{singleviewbg}}\n"
               f"\\multicolumn{{7}}{{l}}{{\\itshape "
               f"${esc(name)}=\\{{{qset}\\}}$:\\ "
               f"meet $\\sqcap {esc(name)}=\\mathrm{{DB}}_{{{m}}}$,\\ "
               f"join $\\sqcup {esc(name)}=\\mathrm{{DB}}_{{{j}}}$}} \\\\")
        print(("\\midrule\n" if gi else "") + hdr)
        for role, node in (("normalized", 0), ("meet $\\sqcap$", m),
                           ("join $\\sqcup$", j)):
            r = next(x for x in rows if x[0] == name and (
                (role.startswith("normalized") and x[2] == "normalized") or
                (role.startswith("meet") and x[2] == "meet") or
                (role.startswith("join") and x[2] == "join")))
            red = r[5]
            per = r[4]
            cells = " & ".join(
                f"${per[sf][0]:.2f}\\times$ & {fk(per[sf][1])}" for sf in
                [s for s in args.sfs.split(',') if s.strip()])
            print(f"{role} & $\\mathrm{{DB}}_{{{node}}}$ & {red:.2f} & {cells} \\\\")


if __name__ == "__main__":
    main()
