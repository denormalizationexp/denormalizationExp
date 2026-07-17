#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_denorm_experiment_unified.py
================================
One-shot denormalization-lattice experiment that runs BOTH families in a single
invocation, materialized views only (plain views are intentionally not run):

  * the single-view nodes  DB0, DB1, ..., DB15   (DB0 = normalized baseline), and
  * the 7 complementary multi-view pairs, in the order
        7+8, 6+9, 5+10, 4+11, 3+12, 2+13, 1+14.

It reuses run_denorm_experiment_single.py (schema build + the single-view answering
policy) as module `R`, and folds in the complementary-pair machinery that used
to live in the *_mv companion, so only TWO files are needed:
    run_denorm_experiment_single.py   (imported here as R)
    run_denorm_experiment_unified.py   (this file)

New in this version
-------------------
1. Per-query breakdown.  In addition to the TOTAL workload time, we record the
   time of EACH query Q_j on every instance.  This makes the per-query saving
   versus DB0 available one-to-one:  saving_i[j] = q0[j] - qi[j].

2. Subset switches (mix and match freely):
     --dbs      which single-view nodes to run, 0..15  (DB0 is always measured
                as the baseline and always shown).   e.g.  0,1,3,4,6,8,10,13
     --pairs    which complementary pairs to run.     e.g.  7+8,6+9  | all | none
     --queries  which Q_j make up the workload, 1..15. e.g.  1,3,4,6,8,10,13
   Each accepts a comma list, ranges like 1-15, or the keyword `all`
   (`none` is also accepted for --pairs).

Outputs (one .xlsx workbook, three sheets):
    summary           one row per instance: create / workload(total) /
                      total_exec / workload_ratio / workload_saving(total) /
                      redundancy / total_attr_values
    per_query_time    instance x Q_j : the measured time of each query
    per_query_saving  instance x Q_j : q0[j] - qi[j]   (DB0 row is all zeros)

Example (the worked example from the request):
    python run_denorm_experiment_unified.py --sf-sweep 10,100 \
        --dbname sportsdb --user YOUR_USERNAME --password YOUR_PASSWORD \
        --dbs 0,1,3,4,6,8,10,13 --queries 1,3,4,6,8,10,13 --pairs none \
        --optimize B
"""

import argparse
import time

import run_denorm_experiment_single as R   # schema build + single-view answering


VIEW_TYPE = "materialized"               # this unified runner is materialized-only


# ========================================================================== #
#  PART P — complementary-pair machinery (folded in from the _mv companion)   #
# ========================================================================== #

# Primary key of each relation alias, in the column names a view exposes
# (component_sql exposes column c of alias al as  al_c).  Two views that SHARE a
# relation must be equated on that relation's FULL primary key.
PK_COLS = {
    "p":  ["name"],
    "s":  ["date", "h_name", "p_name"],
    "g":  ["date", "h_name"],
    "th": ["name"],
    "ta": ["name"],
}

# The 7 complementary-pair cases, by lattice-figure DB index.
PAIR_CASES = [
    ("(i)",   (1, 14)),
    ("(ii)",  (2, 13)),
    ("(iii)", (3, 12)),
    ("(iv)",  (4, 11)),
    ("(v)",   (5, 10)),
    ("(vi)",  (6, 9)),
    ("(vii)", (7, 8)),
]

# relation grain, finest first: a view's primary key is the PK of the finest
# relation it contains (stats > game > player/team).
GRAIN_ORDER = ["s", "g", "p", "th", "ta"]


def view_pk_cols(valiases):
    """PK columns (in view column naming al_c) of a view = PK of its finest rel."""
    for al in GRAIN_ORDER:
        if al in valiases:
            return [f"{al}_{c}" for c in PK_COLS[al]]
    return []


def build_case_views(cur, nodes, comps_of, db_indices, view_type=VIEW_TYPE):
    """
    Materialize the views of ALL nodes in `db_indices` simultaneously, each with
    a primary-key index (so the optimizer can join two views on a shared relation
    via that index).  Index-build time is included in create_time_s.
    Returns (view_comps, create_time_s, n_views).
    """
    kw = "MATERIALIZED VIEW" if view_type == "materialized" else "VIEW"
    R.drop_all_views(cur)
    view_comps = []
    create_s = 0.0
    for i in db_indices:
        node = nodes[i]
        for c_idx, (aliases, atoms_in) in enumerate(comps_of[i]):
            mv = f"mv_{R.node_key(node)}_{c_idx}"
            sql = R.component_sql(aliases, atoms_in)
            t = time.perf_counter()
            cur.execute(f"CREATE {kw} {mv} AS {sql}")
            pk = view_pk_cols(aliases)
            if pk and view_type == "materialized":
                cur.execute(f"CREATE INDEX ON {mv} ({', '.join(pk)})")
                if "s" in aliases and "g" in aliases:
                    gcols = [f"g_{c}" for c in PK_COLS["g"]]
                    cur.execute(f"CREATE INDEX ON {mv} ({', '.join(gcols)})")
                cur.execute(f"VACUUM ANALYZE {mv}")
            create_s += time.perf_counter() - t
            view_comps.append((set(aliases), frozenset(atoms_in), mv))
    return view_comps, create_s, len(view_comps)


def plan_views_only(cj_atoms, view_comps):
    """Answer a query (atom set cj_atoms) using ONLY the materialized views."""
    cj_atoms = frozenset(cj_atoms)
    for vc in view_comps:                         # single view subsumes the query
        if cj_atoms <= vc[1]:
            return dict(views=[vc], single=True)
    chosen, covered = [], set()                   # else combine views (largest 1st)
    for vc in sorted(view_comps, key=lambda v: len(v[1]), reverse=True):
        if vc[1] & cj_atoms and not (set(vc[1]) <= covered):
            chosen.append(vc)
            covered |= set(vc[1])
        if cj_atoms <= covered:
            break
    if cj_atoms <= covered:
        return dict(views=chosen, single=False)
    return None


def _needed_aliases(views, cj_aliases):
    """Per view, the relation-aliases it must still expose (projection + join
    keys).  Used by strategy B to decide which non-stats columns to keep."""
    cj = set(cj_aliases)
    first_owner = {}
    for k, (valiases, _va, _mv) in enumerate(views):
        for al in valiases:
            first_owner.setdefault(al, k)
    needed = {k: set() for k in range(len(views))}
    for k, (valiases, _va, _mv) in enumerate(views):
        needed[k] |= (set(valiases) & cj)
    from collections import Counter
    holders = Counter()
    for valiases, _va, _mv in views:
        for al in set(valiases):
            holders[al] += 1
    for k, (valiases, _va, _mv) in enumerate(views):
        for al in valiases:
            if holders[al] >= 2:
                needed[k].add(al)
    allatoms = set()
    for _va, vatoms, _mv in views:
        allatoms |= set(vatoms)
    for a in allatoms:
        x, y = R.ATOMS[a][0]
        ox, oy = first_owner.get(x), first_owner.get(y)
        if ox is not None and oy is not None and ox != oy:
            needed[ox].add(x); needed[ox].add(y) if y in views[ox][0] else None
            needed[oy].add(y); needed[oy].add(x) if x in views[oy][0] else None
    return needed


def _view_exposes(views, k, al, optimize, query_has_s, needed=None):
    """does view k still expose relation alias `al` after a possible B-shrink?"""
    valiases = views[k][0]
    if al not in valiases:
        return False
    if optimize and not query_has_s and "s" in valiases:
        if al == "s":
            return False
        if needed is not None and al not in needed[k]:
            return False
    return True


def plan_atoms(views):
    s = set()
    for _va, vatoms, _mv in views:
        s |= set(vatoms)
    return s


def views_only_sql(cj_aliases, plan, optimize=False):
    """Counting SQL that answers the query using only views (strategy A or B)."""
    views = plan["views"]

    if plan.get("single"):
        valiases, _vatoms, mv = views[0]
        proj = ", ".join(f"{al}_{c}"
                         for al in sorted(cj_aliases,
                                          key=lambda x: R.ALIAS_ORDER.index(x))
                         for c in R.cols(al))
        distinct = "DISTINCT " if (("s" in valiases) and
                                   ("s" not in cj_aliases)) else ""
        return f"SELECT count(*) FROM (SELECT {distinct}{proj} FROM {mv}) q"

    query_has_s = "s" in cj_aliases
    needed = _needed_aliases(views, cj_aliases)

    owner = {}
    vtag_of = {}
    from_parts = []
    for k, (valiases, _va, mv) in enumerate(views):
        vtag = f"v{k}"
        vtag_of[k] = vtag
        if (optimize and not query_has_s and "s" in valiases):
            keep_aliases = [al for al in
                            sorted(valiases, key=lambda x: R.ALIAS_ORDER.index(x))
                            if al != "s" and al in needed[k]]
            keep_cols = ", ".join(f"{al}_{c}" for al in keep_aliases
                                  for c in R.cols(al))
            from_parts.append(
                f"(SELECT DISTINCT {keep_cols} FROM {mv}) {vtag}")
            for al in keep_aliases:
                owner.setdefault(al, vtag)
        else:
            from_parts.append(f"(SELECT * FROM {mv}) {vtag}")
            for al in valiases:
                owner.setdefault(al, vtag)

    preds = []
    for al in PK_COLS:
        holders = [vt for k, vt in vtag_of.items()
                   if al in views[k][0] and _view_exposes(views, k, al, optimize,
                                                          query_has_s, needed)]
        if len(holders) >= 2:
            base = holders[0]
            for other in holders[1:]:
                for c in PK_COLS[al]:
                    preds.append(f"{base}.{al}_{c} = {other}.{al}_{c}")

    def colref(al, c):
        return f"{owner[al]}.{al}_{c}"

    for a in plan_atoms(views):
        x, y = R.ATOMS[a][0]
        # Skip an atom whose two endpoints both live inside a SINGLE chosen view:
        # that view already materializes the join, so re-emitting it as a cross-
        # view predicate is redundant (the views are equated on their shared
        # relation's PK).  The `owner` map records only one view per relation, so
        # without this guard a shared hub relation (e.g. `g`, present in both
        # views) makes a within-view dimension atom (e.g. Ga = g-ta, internal to
        # the narrow view) look cross-view.  The redundant equality is harmless
        # for a key column (folds into the PK equivalence class) but forces a
        # residual filter + pathological Memoize for a non-key column, so its
        # omission both fixes correctness of the plan intent and removes a large
        # home/away asymmetry.  Mirrors plan_sql's internal-atom skip.
        if any({x, y} <= set(v[0]) for v in views):
            continue
        if owner.get(x) and owner.get(y) and owner[x] != owner[y]:
            if a == "PS":
                preds.append(f"{colref('p','name')} = {colref('s','p_name')}")
            elif a == "SG":
                preds.append(f"{colref('s','date')} = {colref('g','date')} AND "
                             f"{colref('s','h_name')} = {colref('g','h_name')}")
            elif a == "Gh":
                preds.append(f"{colref('g','h_name')} = {colref('th','name')}")
            elif a == "Ga":
                preds.append(f"{colref('g','a_name')} = {colref('ta','name')}")

    def pred_view_tags(pred):
        tags = []
        for tok in pred.replace(" AND ", " ").split():
            if "." in tok:
                t = tok.split(".")[0]
                if t.startswith("v") and t[1:].isdigit() and t not in tags:
                    tags.append(t)
        return tags

    pair_preds = {}
    self_preds = []
    for p in preds:
        for atomic in p.split(" AND "):
            tags = pred_view_tags(atomic)
            if len(tags) == 2:
                pair_preds.setdefault(frozenset(tags), []).append(atomic)
            else:
                self_preds.append(atomic)

    ordered_tags = [f"v{k}" for k in range(len(views))]
    in_from = {ordered_tags[0]}
    from_sql = from_parts[0]
    leftover_pairs = dict(pair_preds)
    for vt in ordered_tags[1:]:
        on_preds = []
        for key in list(leftover_pairs.keys()):
            if vt in key and (key - {vt}) <= in_from:
                on_preds.extend(leftover_pairs.pop(key))
        part = from_parts[ordered_tags.index(vt)]
        if on_preds:
            from_sql += f" JOIN {part} ON " + " AND ".join(on_preds)
        else:
            from_sql += f" CROSS JOIN {part}"
        in_from.add(vt)

    where_preds = list(self_preds)
    for key, plist in leftover_pairs.items():
        where_preds.extend(plist)

    proj = ", ".join(colref(al, c)
                     for al in sorted(cj_aliases,
                                      key=lambda x: R.ALIAS_ORDER.index(x))
                     for c in R.cols(al))
    avail_aliases = set(owner)
    distinct = "DISTINCT " if (("s" in avail_aliases) and
                               ("s" not in cj_aliases)) else ""
    inner = f"SELECT {distinct}{proj} FROM {from_sql}"
    if where_preds:
        inner += " WHERE " + " AND ".join(where_preds)
    return f"SELECT count(*) FROM ({inner}) q"


def base_attr_values(c):
    """Total stored attribute values (rows x cols) of DB0, over its 4 base tables.
    Team is one physical table (th/ta are roles), counted once.  Returns a dict
    keyed by PHYSICAL table name so the single-view leftover logic can use it."""
    return {"player": c["players"] * len(R.cols("p")),
            "team":   c["teams"]   * len(R.cols("th")),
            "game":   c["games"]   * len(R.cols("g")),
            "stats":  c["stats"]   * len(R.cols("s"))}


def case_redundancy(cur, view_comps, base_total):
    """Redundancy of a complementary-pair case (REPLACEMENT semantics): the pair
    covers all four atoms, so every base relation is absorbed into some view and
    the stored data is exactly the views.  total_attr = sum over views of
    rows x cols.  Returns (total_attr, total_attr / base_total)."""
    total_attr = 0
    for aliases, _atoms, mv in view_comps:
        cur.execute(f"SELECT count(*) FROM {mv}")
        nrows = cur.fetchone()[0]
        ncols = sum(len(R.cols(al)) for al in aliases)
        total_attr += nrows * ncols
    return total_attr, (total_attr / base_total if base_total else 0.0)


# ========================================================================== #
#  PART Q — per-query breakdown workloads                                     #
# ========================================================================== #

def single_workload_breakdown(cur, view_comps, comps_of, queries, repeats, warmup):
    """
    Per-query time of the workload on a single-view DBi under R's view-reuse
    policy.  Returns {j: time_of_Qj} for j in `queries`.  Qj's time is the sum
    over its (possibly several) connected components.  For DB0 pass view_comps=[]
    so every query falls to the base path (the normalized baseline).
    """
    per_q = {}
    for j in queries:
        tj = 0.0
        for cj_aliases, cj_atoms_in in comps_of[j]:
            cj_atoms = frozenset(cj_atoms_in)
            plan = R.plan_query(cj_atoms, view_comps)
            kind = plan["kind"]
            if kind == "view_exact":
                tj += R.view_answer_cost(cur, cj_aliases, plan["views"][0],
                                         True, repeats, warmup)
            elif kind == "view_subset":
                tj += R.view_answer_cost(cur, cj_aliases, plan["views"][0],
                                         False, repeats, warmup)
            elif kind == "reuse":
                t, _sql = R.plan_answer_cost(cur, cj_aliases, plan,
                                             repeats, warmup)
                tj += t
            else:  # base
                tj += R.base_comp_cost(cur, cj_aliases, cj_atoms_in,
                                       repeats, warmup)
        per_q[j] = tj
    return per_q


def pairs_workload_breakdown(cur, view_comps, comps_of, queries, repeats, warmup,
                             optimize=False):
    """Per-query time of the workload answered using ONLY the views.
    Returns {j: time_of_Qj} for j in `queries`."""
    per_q = {}
    for j in queries:
        tj = 0.0
        for cj_aliases, cj_atoms_in in comps_of[j]:
            plan = plan_views_only(frozenset(cj_atoms_in), view_comps)
            if plan is None:
                raise RuntimeError(
                    f"query {sorted(cj_atoms_in)} not coverable by views alone")
            sql = views_only_sql(cj_aliases, plan, optimize)
            tj += R.agg_time(cur, sql, repeats, warmup)
        per_q[j] = tj
    return per_q


# ========================================================================== #
#  PART R — per-instance runners (with breakdown)                             #
# ========================================================================== #

def _finalize(rec, q0, queries):
    """attach total workload, ratio, total saving and per-query saving."""
    total0 = sum(q0[j] for j in queries)
    wl = sum(rec["per_q"][j] for j in queries)
    rec["workload_s"] = wl
    rec["total_s"] = rec["create_s"] + wl
    rec["wl_ratio"] = (wl / total0) if total0 else 0.0
    rec["saving_s"] = total0 - wl
    rec["per_q_saving"] = {j: q0[j] - rec["per_q"][j] for j in queries}
    return rec


def run_db0(cur, comps_of, queries, base_dims, repeats, warmup):
    """DB0 baseline: per-query times on the normalized base tables (no views)."""
    q0 = single_workload_breakdown(cur, [], comps_of, queries, repeats, warmup)
    base_total = sum(base_dims.values())
    rec = dict(kind="base", order=0, label="DB0", node="normalized (no views)",
               total_attr=base_total, redundancy=1.0, create_s=0.0, per_q=q0)
    return _finalize(rec, q0, queries), q0


def run_single(cur, i, nodes, comps_of, q0, queries, base_dims, repeats, warmup):
    """Single-view node DBi: build its view(s), measure per-query times, redundancy."""
    node = nodes[i]
    base_total = sum(base_dims.values())
    R.drop_all_views(cur)
    create_s, view_attr = 0.0, 0
    view_comps, involved = [], set()
    for c_idx, (aliases, atoms_in) in enumerate(comps_of[i]):
        mv = f"mv_{R.node_key(node)}_{c_idx}"
        sql = R.component_sql(aliases, atoms_in)
        t = time.perf_counter()
        cur.execute(f"CREATE MATERIALIZED VIEW {mv} AS {sql}")
        # PK index + VACUUM ANALYZE on the view, same as the pair experiment
        pk = view_pk_cols(aliases)
        if pk:
            cur.execute(f"CREATE INDEX ON {mv} ({', '.join(pk)})")
            if "s" in aliases and "g" in aliases:
                gcols = [f"g_{c}" for c in PK_COLS["g"]]
                cur.execute(f"CREATE INDEX ON {mv} ({', '.join(gcols)})")
            cur.execute(f"VACUUM ANALYZE {mv}")
        create_s += time.perf_counter() - t
        view_comps.append((set(aliases), frozenset(atoms_in), mv))
        cur.execute(f"SELECT count(*) FROM {mv}")
        view_attr += cur.fetchone()[0] * len(R.proj_cols(aliases))
        involved |= {R.TABLES[al][0] for al in aliases}
    per_q = single_workload_breakdown(cur, view_comps, comps_of, queries,
                                      repeats, warmup)
    R.drop_all_views(cur)
    # redundancy (replacement semantics): merged view(s) + base tables not absorbed
    leftover_attr = sum(v for t, v in base_dims.items() if t not in involved)
    total_attr = view_attr + leftover_attr
    rec = dict(kind="single", order=i, label=str(i),
               node=R.node_label(node), total_attr=total_attr,
               redundancy=total_attr / base_total, create_s=create_s, per_q=per_q)
    return _finalize(rec, q0, queries)


def run_pair(cur, name, dbs, nodes, comps_of, q0, queries, base_total,
             repeats, warmup, optimize):
    """Complementary pair: build both nodes' views, measure per-query times,
    redundancy.  Every query is answered from the views alone."""
    view_comps, create_s, n_views = build_case_views(cur, nodes, comps_of, dbs)
    per_q = pairs_workload_breakdown(cur, view_comps, comps_of, queries,
                                     repeats, warmup, optimize)
    total_attr, redundancy = case_redundancy(cur, view_comps, base_total)
    R.drop_all_views(cur)
    label = "+".join(str(i) for i in dbs)                      # e.g. "7+8"
    node = " | ".join(R.node_label(nodes[i]) for i in dbs)    # "Gh Ga | PS SG"
    rec = dict(kind="pair", order=100 + dbs[0], label=label, node=node,
               total_attr=total_attr, redundancy=redundancy,
               create_s=create_s, n_views=n_views, per_q=per_q)
    return _finalize(rec, q0, queries)


# ========================================================================== #
#  PART S — subset parsing                                                    #
# ========================================================================== #

def parse_index_set(spec, lo, hi):
    """'all'/''/None -> [lo..hi];  '0,1,3'  / '1-5,8,10-12' -> sorted unique list,
    clamped to [lo, hi]."""
    if spec is None or str(spec).strip().lower() in ("all", "*", ""):
        return list(range(lo, hi + 1))
    out = []
    for tok in str(spec).split(","):
        tok = tok.strip()
        if not tok:
            continue
        if "-" in tok:
            a, b = tok.split("-", 1)
            out.extend(range(int(a), int(b) + 1))
        else:
            out.append(int(tok))
    return sorted({v for v in out if lo <= v <= hi})


def parse_pairs(spec):
    """'all'/''/None -> all 7 pairs in run order (7+8 first);  'none' -> [];
    else a comma list of 'a+b' labels (order/spacing-insensitive)."""
    canon = list(reversed(PAIR_CASES))           # 7+8, 6+9, ..., 1+14
    s = (spec or "").strip().lower()
    if s in ("all", "*", ""):
        return canon
    if s in ("none", "-"):
        return []
    want = {tok.strip().replace(" ", "") for tok in str(spec).split(",")
            if tok.strip()}
    out = []
    for nm, dbs in canon:
        keys = {f"{dbs[0]}+{dbs[1]}", f"{dbs[1]}+{dbs[0]}", nm}
        if keys & want:
            out.append((nm, dbs))
    return out


# ========================================================================== #
#  PART T — workbook output                                                   #
# ========================================================================== #

def write_xlsx(path, results_by_sf, queries):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hdr = Font(name="Arial", bold=True)
    body = Font(name="Arial")
    fill = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="center")
    qcols = sorted(queries)

    wb = Workbook()

    # ---------- sheet 1: summary ----------
    ws = wb.active
    ws.title = "summary"
    headers = ["sf", "i", "node", "total_attr_values", "redundancy_vs_db0",
               "create_time_s", "workload_time_s", "total_exec_time_s",
               "workload_ratio_vs_db0", "workload_saving_vs_db0_s"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(1, ci, h); c.font = hdr; c.fill = fill; c.alignment = center
    r = 2
    for sf in sorted(results_by_sf):
        for rec in results_by_sf[sf]:
            vals = [sf, rec["label"], rec["node"], rec["total_attr"],
                    rec["redundancy"], rec["create_s"], rec["workload_s"],
                    rec["total_s"], rec["wl_ratio"], rec["saving_s"]]
            for ci, v in enumerate(vals, start=1):
                ws.cell(r, ci, v).font = body
            for col, fmt in ((4, "#,##0"), (5, "0.0000"), (6, "0.000000"),
                             (7, "0.000000"), (8, "0.000000"), (9, "0.0000"),
                             (10, "0.000000")):
                ws.cell(r, col).number_format = fmt
            r += 1
    for idx, w in enumerate([6, 9, 24, 18, 18, 16, 17, 18, 21, 24], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "D2"

    # ---------- sheets 2 & 3: per-query time and per-query saving ----------
    def per_query_sheet(title, field):
        wsx = wb.create_sheet(title)
        head = ["sf", "i", "node"] + [f"Q{j}" for j in qcols]
        for ci, h in enumerate(head, start=1):
            c = wsx.cell(1, ci, h); c.font = hdr; c.fill = fill; c.alignment = center
        rr = 2
        for sf in sorted(results_by_sf):
            for rec in results_by_sf[sf]:
                wsx.cell(rr, 1, sf).font = body
                wsx.cell(rr, 2, rec["label"]).font = body
                wsx.cell(rr, 3, rec["node"]).font = body
                for k, j in enumerate(qcols):
                    c = wsx.cell(rr, 4 + k, rec[field][j]); c.font = body
                    c.number_format = "0.000000"
                rr += 1
        for idx in range(1, 3 + len(qcols) + 1):
            w = (6 if idx == 1 else 9 if idx == 2 else 24 if idx == 3 else 12)
            wsx.column_dimensions[get_column_letter(idx)].width = w
        wsx.freeze_panes = "D2"

    per_query_sheet("per_query_time", "per_q")
    per_query_sheet("per_query_saving", "per_q_saving")

    wb.save(path)


# ========================================================================== #
#  Main                                                                       #
# ========================================================================== #
def main():
    ap = argparse.ArgumentParser(
        description="Unified single-view + complementary-pair denormalization "
                    "experiment (materialized views), with per-query breakdown "
                    "and DB/query subset switches.")
    # connection
    ap.add_argument("--dbname", default="sportsdb")
    ap.add_argument("--host", default="localhost")
    ap.add_argument("--port", default=5432, type=int)
    ap.add_argument("--user", default="", help="PostgreSQL username.")
    ap.add_argument("--password", default="", help="PostgreSQL password.")
    # scale-factor sweep + build params
    ap.add_argument("--sf-sweep", default="10,100",
                    help="comma-separated scale factors (seasons), e.g. 1,5,10")
    ap.add_argument("--teams", type=int, default=R.TEAMS)
    ap.add_argument("--games-per-season", type=int, default=R.GAMES_PER_SEASON)
    ap.add_argument("--players-per-season", type=int, default=R.PLAYERS_PER_SEASON)
    ap.add_argument("--turnover", type=float, default=R.TURNOVER)
    ap.add_argument("--per-game", type=int, default=R.STATS_PER_GAME)
    ap.add_argument("--seed", type=int, default=42)
    # subset switches
    ap.add_argument("--dbs", default="0,1,3,4,6,8,10,13",
                    help="single-view nodes to run, 0..15 (DB0 is always the "
                         "baseline). e.g. 0,1,3,4,6,8,10,13  or  0-15  or  all")
    ap.add_argument("--pairs", default="none",
                    help="complementary pairs to run: all | none | a list like "
                         "7+8,6+9,1+14")
    ap.add_argument("--queries", default="1,3,4,6,8,10,13",
                    help="which Q_j make up the workload, 1..15. "
                         "e.g. 1,3,4,6,8,10,13  or  1-15  or  all")
    # experiment params
    ap.add_argument("--query-repeats", type=int, default=40)
    ap.add_argument("--agg", choices=["median", "mean"], default="mean",
                    help="aggregate repeated query timings by median or mean")
    ap.add_argument("--warmup", type=int, default=2)
    ap.add_argument("--optimize", choices=["A", "B"], default="B",
                    help="pair answering: A = direct view stitching; "
                         "B = pre-shrink stats-bearing views before joining. "
                         "(single-view nodes always use R's reuse policy.)")
    ap.add_argument("--work-mem", default="256MB")
    ap.add_argument("--max-parallel", type=int, default=0)
    ap.add_argument("--xlsx", default="denorm_unified_results_partial.xlsx")
    ap.add_argument("--queries-file", default="workload_queries_partial.sql")
    ap.add_argument("--db0-workload", default="",
                    help="optional precomputed DB0 TOTAL workload per sf as "
                         "'sf:seconds' pairs; only used for the console ratio "
                         "print, NOT for the per-query saving (which always "
                         "needs the measured per-query DB0 times).")
    args = ap.parse_args()

    R.AGG = R.AGG_FUNCS[args.agg]
    sfs = [float(x) for x in args.sf_sweep.split(",") if x.strip()]

    single_dbs = [i for i in parse_index_set(args.dbs, 0, 15) if i != 0]
    pairs = parse_pairs(args.pairs)
    queries = parse_index_set(args.queries, 1, 15)
    if not queries:
        raise SystemExit("no queries selected (--queries)")

    conn = R.connect_or_create(args)
    conn.autocommit = True
    cur = conn.cursor()
    if args.work_mem:
        cur.execute(f"SET work_mem = '{args.work_mem}'")
        print(f"[cfg] work_mem = {args.work_mem}")
    if args.max_parallel is not None:
        cur.execute(f"SET max_parallel_workers_per_gather = {args.max_parallel}")
        print(f"[cfg] max_parallel_workers_per_gather = {args.max_parallel}")

    nodes = R.all_nodes()
    comps_of = {i: R.components(n) for i, n in enumerate(nodes)}
    R.save_queries(args.queries_file, nodes, comps_of)
    print(f"[ok] wrote {args.queries_file} (Q1..Q15)")
    print(f"[cfg] single-view DBs = [0] + {single_dbs}")
    print(f"[cfg] pairs = {[f'{a}+{b}' for _n,(a,b) in pairs]}")
    print(f"[cfg] workload queries = {queries}")
    print(f"[cfg] pair strategy = {args.optimize}")

    results_by_sf = {}
    for sf in sfs:
        print(f"\n############### scale factor sf = {sf} ###############")
        c = R.build_database(conn, sf, args.teams, args.games_per_season,
                             args.players_per_season, args.turnover,
                             args.per_game, args.seed)
        print(f"[build] teams={c['teams']} players={c['players']} "
              f"games={c['games']} stats={c['stats']}")
        base_dims = base_attr_values(c)
        base_total = sum(base_dims.values())
        print(f"[db0]  base attribute values = {base_total:,}")
        cur.execute("ANALYZE")
        R.drop_all_views(cur)
        R._base_comp_cost.clear()            # data changed -> invalidate caches

        # ---- DB0 baseline (per-query) ----
        db0_rec, q0 = run_db0(cur, comps_of, queries, base_dims,
                              args.query_repeats, args.warmup)
        total0 = sum(q0[j] for j in queries)
        print(f"[db0]  workload(DB0) = {total0*1000:.1f} ms over {len(queries)} "
              f"queries (repeats={args.query_repeats}, agg={args.agg})")

        recs = [db0_rec]

        # ---- single-view nodes (ascending) ----
        for i in single_dbs:
            rec = run_single(cur, i, nodes, comps_of, q0, queries, base_dims,
                             args.query_repeats, args.warmup)
            print(f"  DB{i:<2} {rec['node']:<16} create={rec['create_s']*1000:8.1f}ms "
                  f"workload={rec['workload_s']*1000:8.1f}ms "
                  f"(x{rec['wl_ratio']:4.2f}) saving={rec['saving_s']*1000:8.1f}ms "
                  f"redund={rec['redundancy']:5.2f}")
            recs.append(rec)

        # ---- complementary pairs (7+8, 6+9, ..., 1+14) ----
        for nm, dbs in pairs:
            rec = run_pair(cur, nm, dbs, nodes, comps_of, q0, queries, base_total,
                           args.query_repeats, args.warmup,
                           optimize=(args.optimize == "B"))
            print(f"  {rec['label']:<6} {rec['node']:<20} "
                  f"views={rec.get('n_views','?')} "
                  f"create={rec['create_s']*1000:8.1f}ms "
                  f"workload={rec['workload_s']*1000:8.1f}ms "
                  f"(x{rec['wl_ratio']:4.2f}) saving={rec['saving_s']*1000:8.1f}ms "
                  f"redund={rec['redundancy']:5.2f}")
            recs.append(rec)

        results_by_sf[sf] = recs
    conn.close()

    write_xlsx(args.xlsx, results_by_sf, queries)
    print(f"\n[ok] wrote {args.xlsx}  (sheets: summary, per_query_time, "
          f"per_query_saving)")


if __name__ == "__main__":
    main()
