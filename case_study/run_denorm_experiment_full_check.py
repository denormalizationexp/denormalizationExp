#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_denorm_experiment_full.py
=============================
Full case-study experiment runner.

This script is intended to cover the useful surface of both:

  * run_denorm_experiment_single.py
  * run_denorm_experiment_unified.py

It reuses run_denorm_experiment_single.py as module R for schema generation,
lattice metadata, query generation, and the single-view reuse policy.

Features:
  * DB0 + selectable single-view nodes DB1..DB15.
  * Optional complementary materialized-view pairs: 7+8, 6+9, ..., 1+14.
  * Selectable workload queries Q1..Q15.
  * materialized, plain, or both single-view modes.
  * Per-query timing and per-query saving output.
  * A long-format summary plus a single-style summary sheet.

Important notes:
  * Pair experiments are materialized-view experiments only. If --view-type both
    is used, plain single-view rows are produced, and pair rows are produced only
    for materialized views.
  * The script rebuilds the sports tables for every scale factor by calling
    R.build_database(), which drops and recreates team/player/game/stats.

Examples:
  # Full materialized experiment: all DB nodes, all queries, all pairs.
  python run_denorm_experiment_full.py --dbname sportsdb \
      --user YOUR_USERNAME --password YOUR_PASSWORD

  # Reproduce the single-runner surface more closely: all queries/nodes,
  # materialized + plain, no pairs, and no materialized-view indexes.
  python run_denorm_experiment_full.py --view-type both --pairs none \
      --mv-indexes none

  # Generate a workbook usable by analyze_meet_join.py.
  python run_denorm_experiment_full.py --view-type materialized --queries all \
      --dbs all --pairs none
"""

import argparse
import time
from pathlib import Path

import run_denorm_experiment_single as R


SCRIPT_DIR = Path(__file__).resolve().parent


# -------------------------------------------------------------------------- #
# Complementary-pair metadata                                                #
# -------------------------------------------------------------------------- #

PK_COLS = {
    "p": ["name"],
    "s": ["date", "h_name", "p_name"],
    "g": ["date", "h_name"],
    "th": ["name"],
    "ta": ["name"],
}

# Logical source-table keys as they are exposed by a materialized view.  These
# must remain ordinary (non-UNIQUE) indexes: a join can repeat the source PK
# values in the materialized result.
MV_KEY_COLS = {
    "p": [
        ("name",),  # player PK
    ],
    "s": [
        ("date", "h_name", "p_name"),  # stats PK
        ("p_name",),  # stats -> player FK
        ("date", "h_name"),  # stats -> game FK
    ],
    "g": [
        ("date", "h_name"),  # game PK
        ("h_name",),  # game -> home team FK
        ("a_name",),  # game -> away team FK
    ],
    "th": [
        ("name",),  # home-team PK
    ],
    "ta": [
        ("name",),  # away-team PK
    ],
}

# PostgreSQL creates indexes for primary keys, but not for foreign keys.  The
# stats -> game FK (date, h_name) is already covered by the left prefix of the
# stats PK index (date, h_name, p_name), so only the three missing supporting
# indexes need to be created explicitly on DB0.
BASE_FK_INDEXES = [
    ("idx_stats_p_name", "stats", ("p_name",)),
    ("idx_game_h_name", "game", ("h_name",)),
    ("idx_game_a_name", "game", ("a_name",)),
]

BASE_TABLES = ("team", "player", "game", "stats")

PAIR_CASES = [
    ("(i)", (1, 14)),
    ("(ii)", (2, 13)),
    ("(iii)", (3, 12)),
    ("(iv)", (4, 11)),
    ("(v)", (5, 10)),
    ("(vi)", (6, 9)),
    ("(vii)", (7, 8)),
]

GRAIN_ORDER = ["s", "g", "p", "th", "ta"]


def view_index_cols(valiases):
    """Return non-redundant logical PK/FK index columns present in an MV."""
    candidates = []
    for al in GRAIN_ORDER:
        if al not in valiases:
            continue
        for key_cols in MV_KEY_COLS[al]:
            cols = tuple(f"{al}_{c}" for c in key_cols)
            if cols not in candidates:
                candidates.append(cols)

    # A multi-column B-tree also covers its left prefixes.  For example, the
    # stats PK index (s_date, s_h_name, s_p_name) already covers the composite
    # FK (s_date, s_h_name), so do not create a duplicate physical index.
    return [
        cols
        for cols in candidates
        if not any(
            len(cols) < len(other) and other[:len(cols)] == cols
            for other in candidates
        )
    ]


def should_index(view_type, mv_indexes, kind):
    if view_type != "materialized":
        return False
    if mv_indexes == "all":
        return True
    if mv_indexes == "single":
        return kind == "single"
    if mv_indexes == "pairs":
        return kind == "pair"
    return False


def create_view_indexes(cur, mv, aliases):
    index_cols = view_index_cols(aliases)
    if not index_cols:
        return
    for cols in index_cols:
        cur.execute(f"CREATE INDEX ON {mv} ({', '.join(cols)})")


def prepare_materialized_view(cur, mv, aliases, with_indexes):
    """Create optional MV indexes, then collect comparable stats every time."""
    if with_indexes:
        create_view_indexes(cur, mv, aliases)
    # Run this for indexed and non-indexed MVs alike.  Otherwise an index-mode
    # comparison also changes planner statistics and visibility-map state.
    cur.execute(f"VACUUM ANALYZE {mv}")


def prepare_base_tables(cur):
    """Give DB0 the same logical PK/FK index coverage and maintenance state."""
    for index_name, table, cols in BASE_FK_INDEXES:
        cur.execute(
            f"CREATE INDEX {index_name} ON {table} ({', '.join(cols)})"
        )
    for table in BASE_TABLES:
        cur.execute(f"VACUUM ANALYZE {table}")


def build_case_views(cur, nodes, comps_of, db_indices, view_type, mv_indexes):
    kw = "MATERIALIZED VIEW" if view_type == "materialized" else "VIEW"
    R.drop_all_views(cur)
    view_comps = []
    create_s = 0.0
    for i in db_indices:
        node = nodes[i]
        for c_idx, (aliases, atoms_in) in enumerate(comps_of[i]):
            mv = f"mv_{R.node_key(node)}_{c_idx}"
            sql = R.component_sql(aliases, atoms_in)
            t = time.perf_counter()
            cur.execute(f"CREATE {kw} {mv} AS {sql}")
            if view_type == "materialized":
                prepare_materialized_view(
                    cur,
                    mv,
                    aliases,
                    should_index(view_type, mv_indexes, "pair"),
                )
            create_s += time.perf_counter() - t
            view_comps.append((set(aliases), frozenset(atoms_in), mv))
    return view_comps, create_s, len(view_comps)


# -------------------------------------------------------------------------- #
# Pair query planning                                                         #
# -------------------------------------------------------------------------- #

def plan_views_only(cj_atoms, view_comps):
    cj_atoms = frozenset(cj_atoms)
    for vc in view_comps:
        if cj_atoms <= vc[1]:
            return {"views": [vc], "single": True}

    chosen, covered = [], set()
    for vc in sorted(view_comps, key=lambda v: len(v[1]), reverse=True):
        if vc[1] & cj_atoms and not (set(vc[1]) <= covered):
            chosen.append(vc)
            covered |= set(vc[1])
        if cj_atoms <= covered:
            break
    if cj_atoms <= covered:
        return {"views": chosen, "single": False}
    return None


def _needed_aliases(views, cj_aliases):
    cj = set(cj_aliases)
    first_owner = {}
    for k, (valiases, _vatoms, _mv) in enumerate(views):
        for al in valiases:
            first_owner.setdefault(al, k)

    needed = {k: set() for k in range(len(views))}
    for k, (valiases, _vatoms, _mv) in enumerate(views):
        needed[k] |= set(valiases) & cj

    from collections import Counter

    holders = Counter()
    for valiases, _vatoms, _mv in views:
        for al in set(valiases):
            holders[al] += 1
    for k, (valiases, _vatoms, _mv) in enumerate(views):
        for al in valiases:
            if holders[al] >= 2:
                needed[k].add(al)

    allatoms = set()
    for _valiases, vatoms, _mv in views:
        allatoms |= set(vatoms)
    for a in allatoms:
        x, y = R.ATOMS[a][0]
        ox, oy = first_owner.get(x), first_owner.get(y)
        if ox is not None and oy is not None and ox != oy:
            needed[ox].add(x)
            if y in views[ox][0]:
                needed[ox].add(y)
            needed[oy].add(y)
            if x in views[oy][0]:
                needed[oy].add(x)
    return needed


def _view_exposes(views, k, al, optimize, query_has_s, needed=None):
    valiases = views[k][0]
    if al not in valiases:
        return False
    if optimize and not query_has_s and "s" in valiases:
        if al == "s":
            return False
        if needed is not None and al not in needed[k]:
            return False
    return True


def plan_atoms(views):
    atoms = set()
    for _valiases, vatoms, _mv in views:
        atoms |= set(vatoms)
    return atoms


def views_only_sql(cj_aliases, plan, optimize=False):
    views = plan["views"]

    if plan.get("single"):
        valiases, _vatoms, mv = views[0]
        proj = ", ".join(
            f"{al}_{c}"
            for al in sorted(cj_aliases, key=lambda x: R.ALIAS_ORDER.index(x))
            for c in R.cols(al)
        )
        distinct = "DISTINCT " if (("s" in valiases) and ("s" not in cj_aliases)) else ""
        return f"SELECT count(*) FROM (SELECT {distinct}{proj} FROM {mv}) q"

    query_has_s = "s" in cj_aliases
    needed = _needed_aliases(views, cj_aliases)

    owner = {}
    vtag_of = {}
    from_parts = []
    for k, (valiases, _vatoms, mv) in enumerate(views):
        vtag = f"v{k}"
        vtag_of[k] = vtag
        if optimize and not query_has_s and "s" in valiases:
            keep_aliases = [
                al
                for al in sorted(valiases, key=lambda x: R.ALIAS_ORDER.index(x))
                if al != "s" and al in needed[k]
            ]
            keep_cols = ", ".join(f"{al}_{c}" for al in keep_aliases for c in R.cols(al))
            from_parts.append(f"(SELECT DISTINCT {keep_cols} FROM {mv}) {vtag}")
            for al in keep_aliases:
                owner.setdefault(al, vtag)
        else:
            from_parts.append(f"(SELECT * FROM {mv}) {vtag}")
            for al in valiases:
                owner.setdefault(al, vtag)

    preds = []
    for al in PK_COLS:
        holders = [
            vt
            for k, vt in vtag_of.items()
            if al in views[k][0] and _view_exposes(views, k, al, optimize, query_has_s, needed)
        ]
        if len(holders) >= 2:
            base = holders[0]
            for other in holders[1:]:
                for c in PK_COLS[al]:
                    preds.append(f"{base}.{al}_{c} = {other}.{al}_{c}")

    def colref(al, c):
        return f"{owner[al]}.{al}_{c}"

    for a in plan_atoms(views):
        x, y = R.ATOMS[a][0]
        if any({x, y} <= set(v[0]) for v in views):
            continue
        if owner.get(x) and owner.get(y) and owner[x] != owner[y]:
            if a == "PS":
                preds.append(f"{colref('p', 'name')} = {colref('s', 'p_name')}")
            elif a == "SG":
                preds.append(
                    f"{colref('s', 'date')} = {colref('g', 'date')} AND "
                    f"{colref('s', 'h_name')} = {colref('g', 'h_name')}"
                )
            elif a == "Gh":
                preds.append(f"{colref('g', 'h_name')} = {colref('th', 'name')}")
            elif a == "Ga":
                preds.append(f"{colref('g', 'a_name')} = {colref('ta', 'name')}")

    def pred_view_tags(pred):
        tags = []
        for tok in pred.replace(" AND ", " ").split():
            if "." in tok:
                t = tok.split(".")[0]
                if t.startswith("v") and t[1:].isdigit() and t not in tags:
                    tags.append(t)
        return tags

    pair_preds = {}
    self_preds = []
    for pred in preds:
        for atomic in pred.split(" AND "):
            tags = pred_view_tags(atomic)
            if len(tags) == 2:
                pair_preds.setdefault(frozenset(tags), []).append(atomic)
            else:
                self_preds.append(atomic)

    ordered_tags = [f"v{k}" for k in range(len(views))]
    in_from = {ordered_tags[0]}
    from_sql = from_parts[0]
    leftover_pairs = dict(pair_preds)
    for vt in ordered_tags[1:]:
        on_preds = []
        for key in list(leftover_pairs.keys()):
            if vt in key and (key - {vt}) <= in_from:
                on_preds.extend(leftover_pairs.pop(key))
        part = from_parts[ordered_tags.index(vt)]
        if on_preds:
            from_sql += f" JOIN {part} ON " + " AND ".join(on_preds)
        else:
            from_sql += f" CROSS JOIN {part}"
        in_from.add(vt)

    where_preds = list(self_preds)
    for _key, plist in leftover_pairs.items():
        where_preds.extend(plist)

    proj = ", ".join(
        colref(al, c)
        for al in sorted(cj_aliases, key=lambda x: R.ALIAS_ORDER.index(x))
        for c in R.cols(al)
    )
    avail_aliases = set(owner)
    distinct = "DISTINCT " if (("s" in avail_aliases) and ("s" not in cj_aliases)) else ""
    inner = f"SELECT {distinct}{proj} FROM {from_sql}"
    if where_preds:
        inner += " WHERE " + " AND ".join(where_preds)
    return f"SELECT count(*) FROM ({inner}) q"


# -------------------------------------------------------------------------- #
# Workload measurement                                                        #
# -------------------------------------------------------------------------- #

def base_attr_values(c):
    return {
        "player": c["players"] * len(R.cols("p")),
        "team": c["teams"] * len(R.cols("th")),
        "game": c["games"] * len(R.cols("g")),
        "stats": c["stats"] * len(R.cols("s")),
    }


def case_redundancy(cur, view_comps, base_total):
    total_attr = 0
    for aliases, _atoms, mv in view_comps:
        cur.execute(f"SELECT count(*) FROM {mv}")
        nrows = cur.fetchone()[0]
        ncols = sum(len(R.cols(al)) for al in aliases)
        total_attr += nrows * ncols
    return total_attr, (total_attr / base_total if base_total else 0.0)


def single_workload_breakdown(cur, view_comps, comps_of, queries, repeats, warmup):
    per_q = {}
    for j in queries:
        tj = 0.0
        for cj_aliases, cj_atoms_in in comps_of[j]:
            cj_atoms = frozenset(cj_atoms_in)
            plan = R.plan_query(cj_atoms, view_comps)
            kind = plan["kind"]
            if kind == "view_exact":
                tj += R.view_answer_cost(cur, cj_aliases, plan["views"][0], True, repeats, warmup)
            elif kind == "view_subset":
                tj += R.view_answer_cost(cur, cj_aliases, plan["views"][0], False, repeats, warmup)
            elif kind == "reuse":
                t, _sql = R.plan_answer_cost(cur, cj_aliases, plan, repeats, warmup)
                tj += t
            else:
                tj += R.base_comp_cost(cur, cj_aliases, cj_atoms_in, repeats, warmup)
        per_q[j] = tj
    return per_q


def pairs_workload_breakdown(cur, view_comps, comps_of, queries, repeats, warmup, optimize=False):
    per_q = {}
    for j in queries:
        tj = 0.0
        for cj_aliases, cj_atoms_in in comps_of[j]:
            plan = plan_views_only(frozenset(cj_atoms_in), view_comps)
            if plan is None:
                raise RuntimeError(f"query {sorted(cj_atoms_in)} not coverable by views alone")
            sql = views_only_sql(cj_aliases, plan, optimize)
            tj += R.agg_time(cur, sql, repeats, warmup)
        per_q[j] = tj
    return per_q


def finalize_record(rec, q0, queries):
    total0 = sum(q0[j] for j in queries)
    workload = sum(rec["per_q"][j] for j in queries)
    rec["workload_s"] = workload
    rec["total_s"] = rec["create_s"] + workload
    rec["wl_ratio"] = (workload / total0) if total0 else 0.0
    rec["saving_s"] = total0 - workload
    rec["per_q_saving"] = {j: q0[j] - rec["per_q"][j] for j in queries}
    return rec


def make_db0_record(view_type, q0, queries, base_dims):
    base_total = sum(base_dims.values())
    rec = {
        "kind": "base",
        "order": 0,
        "view_type": view_type,
        "label": "DB0",
        "node": "normalized (no views)",
        "total_attr": base_total,
        "redundancy": 1.0,
        "create_s": 0.0,
        "per_q": dict(q0),
    }
    return finalize_record(rec, q0, queries)


def run_single(cur, view_type, i, nodes, comps_of, q0, queries, base_dims,
               repeats, warmup, mv_indexes):
    node = nodes[i]
    kw = "MATERIALIZED VIEW" if view_type == "materialized" else "VIEW"
    base_total = sum(base_dims.values())
    R.drop_all_views(cur)

    create_s, view_attr = 0.0, 0
    view_comps, involved = [], set()
    for c_idx, (aliases, atoms_in) in enumerate(comps_of[i]):
        mv = f"mv_{R.node_key(node)}_{c_idx}"
        sql = R.component_sql(aliases, atoms_in)
        t = time.perf_counter()
        cur.execute(f"CREATE {kw} {mv} AS {sql}")
        if view_type == "materialized":
            prepare_materialized_view(
                cur,
                mv,
                aliases,
                should_index(view_type, mv_indexes, "single"),
            )
        create_s += time.perf_counter() - t

        view_comps.append((set(aliases), frozenset(atoms_in), mv))
        cur.execute(f"SELECT count(*) FROM {mv}")
        view_attr += cur.fetchone()[0] * len(R.proj_cols(aliases))
        involved |= {R.TABLES[al][0] for al in aliases}

    per_q = single_workload_breakdown(cur, view_comps, comps_of, queries, repeats, warmup)
    R.drop_all_views(cur)

    leftover_attr = sum(v for t, v in base_dims.items() if t not in involved)
    total_attr = view_attr + leftover_attr
    rec = {
        "kind": "single",
        "order": i,
        "view_type": view_type,
        "label": str(i),
        "node": R.node_label(node),
        "total_attr": total_attr,
        "redundancy": total_attr / base_total,
        "create_s": create_s,
        "per_q": per_q,
    }
    return finalize_record(rec, q0, queries)


def run_pair(cur, name, dbs, nodes, comps_of, q0, queries, base_total,
             repeats, warmup, optimize, mv_indexes):
    view_comps, create_s, n_views = build_case_views(
        cur, nodes, comps_of, dbs, "materialized", mv_indexes
    )
    per_q = pairs_workload_breakdown(cur, view_comps, comps_of, queries, repeats, warmup, optimize)
    total_attr, redundancy = case_redundancy(cur, view_comps, base_total)
    R.drop_all_views(cur)

    label = "+".join(str(i) for i in dbs)
    node = " | ".join(R.node_label(nodes[i]) for i in dbs)
    rec = {
        "kind": "pair",
        "order": 100 + dbs[0],
        "view_type": "materialized",
        "label": label,
        "node": node,
        "total_attr": total_attr,
        "redundancy": redundancy,
        "create_s": create_s,
        "n_views": n_views,
        "per_q": per_q,
    }
    return finalize_record(rec, q0, queries)


# -------------------------------------------------------------------------- #
# Argument parsing                                                            #
# -------------------------------------------------------------------------- #

def parse_index_set(spec, lo, hi):
    if spec is None or str(spec).strip().lower() in ("all", "*", ""):
        return list(range(lo, hi + 1))
    out = []
    for tok in str(spec).split(","):
        tok = tok.strip()
        if not tok:
            continue
        if "-" in tok:
            a, b = tok.split("-", 1)
            out.extend(range(int(a), int(b) + 1))
        else:
            out.append(int(tok))
    return sorted({v for v in out if lo <= v <= hi})


def parse_pairs(spec):
    canon = list(reversed(PAIR_CASES))  # 7+8, 6+9, ..., 1+14
    s = (spec or "").strip().lower()
    if s in ("all", "*", ""):
        return canon
    if s in ("none", "-"):
        return []

    want = {tok.strip().replace(" ", "") for tok in str(spec).split(",") if tok.strip()}
    out = []
    matched = set()
    for nm, dbs in canon:
        keys = {f"{dbs[0]}+{dbs[1]}", f"{dbs[1]}+{dbs[0]}", nm}
        hit = keys & want
        if hit:
            out.append((nm, dbs))
            matched |= hit
    unknown = want - matched
    if unknown:
        raise SystemExit(f"unknown --pairs value(s): {', '.join(sorted(unknown))}")
    return out


def parse_sfs(spec):
    return [float(x) for x in str(spec).split(",") if x.strip()]


def sf_cell_value(sf):
    return int(sf) if float(sf).is_integer() else sf


def default_path(value, filename):
    if value:
        return Path(value)
    return SCRIPT_DIR / filename


# -------------------------------------------------------------------------- #
# Workbook output                                                             #
# -------------------------------------------------------------------------- #

def write_xlsx(path, records, queries, modes):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl is required for .xlsx output")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    hdr = Font(name="Arial", bold=True)
    body = Font(name="Arial")
    fill = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="center")
    qcols = sorted(queries)

    records = sorted(records, key=lambda r: (float(r["sf"]), r["view_type"], r["order"], r["label"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    headers = [
        "sf", "i", "node", "view_type", "kind", "total_attr_values",
        "redundancy_vs_db0", "create_time_s", "workload_time_s",
        "total_exec_time_s", "workload_ratio_vs_db0",
        "workload_saving_vs_db0_s", "n_views",
    ]
    write_header(ws, headers, hdr, fill, center)
    for row_idx, rec in enumerate(records, start=2):
        vals = [
            sf_cell_value(rec["sf"]), rec["label"], rec["node"],
            rec["view_type"], rec["kind"], rec["total_attr"],
            rec["redundancy"], rec["create_s"], rec["workload_s"],
            rec["total_s"], rec["wl_ratio"], rec["saving_s"],
            rec.get("n_views"),
        ]
        write_row(ws, row_idx, vals, body)
        for col, fmt in (
            (6, "#,##0"), (7, "0.0000"), (8, "0.000000"),
            (9, "0.000000"), (10, "0.000000"), (11, "0.0000"),
            (12, "0.000000"),
        ):
            ws.cell(row_idx, col).number_format = fmt
    set_widths(ws, [6, 9, 26, 14, 12, 18, 18, 16, 17, 18, 21, 24, 10])
    ws.freeze_panes = "D2"

    write_per_query_sheet(wb, "per_query_time", "per_q", records, qcols, hdr, body, fill, center)
    write_per_query_sheet(wb, "per_query_saving", "per_q_saving", records, qcols, hdr, body, fill, center)
    write_single_style_sheet(wb, records, modes, hdr, body, fill, center)

    wb.save(path)


def write_header(ws, headers, hdr, fill, center):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(1, ci, h)
        c.font = hdr
        c.fill = fill
        c.alignment = center


def write_row(ws, row_idx, vals, body):
    for ci, value in enumerate(vals, start=1):
        ws.cell(row_idx, ci, value).font = body


def set_widths(ws, widths):
    from openpyxl.utils import get_column_letter

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_per_query_sheet(wb, title, field, records, qcols, hdr, body, fill, center):
    ws = wb.create_sheet(title)
    headers = ["sf", "i", "node", "view_type", "kind"] + [f"Q{j}" for j in qcols]
    write_header(ws, headers, hdr, fill, center)
    for row_idx, rec in enumerate(records, start=2):
        vals = [
            sf_cell_value(rec["sf"]), rec["label"], rec["node"],
            rec["view_type"], rec["kind"],
        ] + [rec[field][j] for j in qcols]
        write_row(ws, row_idx, vals, body)
        for col in range(6, 6 + len(qcols)):
            ws.cell(row_idx, col).number_format = "0.000000"
    set_widths(ws, [6, 9, 26, 14, 12] + [12] * len(qcols))
    ws.freeze_panes = "F2"


def write_single_style_sheet(wb, records, modes, hdr, body, fill, center):
    ws = wb.create_sheet("single_style_summary")
    single_records = [r for r in records if r["kind"] in ("base", "single")]

    shared = ["sf", "i", "node", "total_attr_values", "redundancy_vs_db0"]
    metrics = ["create_time_s", "workload_time_s", "total_exec_time_s", "workload_ratio_vs_db0"]
    headers = shared[:]
    for mode in modes:
        for metric in metrics:
            headers.append(f"{mode}_{metric}")
    write_header(ws, headers, hdr, fill, center)

    grouped = {}
    for rec in single_records:
        key = (float(rec["sf"]), rec["order"], rec["label"], rec["node"])
        grouped.setdefault(key, {})[rec["view_type"]] = rec

    row_idx = 2
    for key in sorted(grouped, key=lambda x: (x[0], x[1])):
        sf, _order, label, node = key
        by_mode = grouped[key]
        first = next(iter(by_mode.values()))
        vals = [sf_cell_value(sf), label, node, first["total_attr"], first["redundancy"]]
        for mode in modes:
            rec = by_mode.get(mode)
            if rec:
                vals.extend([rec["create_s"], rec["workload_s"], rec["total_s"], rec["wl_ratio"]])
            else:
                vals.extend([None, None, None, None])
        write_row(ws, row_idx, vals, body)
        ws.cell(row_idx, 4).number_format = "#,##0"
        ws.cell(row_idx, 5).number_format = "0.0000"
        for col in range(6, len(vals) + 1):
            ws.cell(row_idx, col).number_format = "0.000000"
        row_idx += 1

    set_widths(ws, [6, 9, 24, 18, 18] + [18] * (len(headers) - 5))
    ws.freeze_panes = "F2"


# -------------------------------------------------------------------------- #
# Main                                                                        #
# -------------------------------------------------------------------------- #

def main():
    ap = argparse.ArgumentParser(
        description="Full sports denormalization experiment runner."
    )
    ap.add_argument("--dbname", default="sportsdb")
    ap.add_argument("--host", default="localhost")
    ap.add_argument("--port", default=5432, type=int)
    ap.add_argument("--user", default="", help="PostgreSQL username.")
    ap.add_argument("--password", default="", help="PostgreSQL password.")

    ap.add_argument("--sf-sweep", default="10,100",
                    help="comma-separated scale factors, e.g. 1,5,10")
    ap.add_argument("--teams", type=int, default=R.TEAMS)
    ap.add_argument("--games-per-season", type=int, default=R.GAMES_PER_SEASON)
    ap.add_argument("--players-per-season", type=int, default=R.PLAYERS_PER_SEASON)
    ap.add_argument("--turnover", type=float, default=R.TURNOVER)
    ap.add_argument("--per-game", type=R.positive_per_game,
                    default=R.STATS_PER_GAME,
                    help="stats rows generated per game; must be at least 1")
    ap.add_argument("--seed", type=int, default=42)

    ap.add_argument("--view-type", choices=["materialized", "plain", "both"],
                    default="materialized")
    ap.add_argument("--dbs", default="all",
                    help="single-view nodes to run, 0..15; DB0 is always included")
    ap.add_argument("--pairs", default="all",
                    help="materialized complementary pairs: all | none | 7+8,6+9,...")
    ap.add_argument("--queries", default="all",
                    help="workload queries, 1..15; accepts all, 1-15, or comma lists")
    ap.add_argument("--query-repeats", type=int, default=40) #40
    ap.add_argument("--agg", choices=["median", "mean"], default="mean")
    ap.add_argument("--warmup", type=int, default=2)    #2
    ap.add_argument("--optimize", choices=["A", "B"], default="B",
                    help="pair answering strategy")
    ap.add_argument("--mv-indexes", choices=["none", "single", "pairs", "all"],
                    default="all",
                    help="which materialized views get logical PK/FK indexes; "
                         "all materialized views receive VACUUM ANALYZE")
    ap.add_argument("--work-mem", default="256MB")
    ap.add_argument("--max-parallel", type=int, default=0)
    ap.add_argument("--xlsx", default=None,
                    help="output workbook; default is case_study/denorm_full_results_with_fk.xlsx")
    ap.add_argument("--queries-file", default=None,
                    help="query SQL output; default is case_study/workload_queries_full.sql")
    args = ap.parse_args()

    R.AGG = R.AGG_FUNCS[args.agg]

    sfs = parse_sfs(args.sf_sweep)
    modes = ["materialized", "plain"] if args.view_type == "both" else [args.view_type]
    single_dbs = [i for i in parse_index_set(args.dbs, 0, 15) if i != 0]
    pairs = parse_pairs(args.pairs)
    queries = parse_index_set(args.queries, 1, 15)
    if not queries:
        raise SystemExit("no queries selected (--queries)")
    if pairs and args.view_type == "plain":
        raise SystemExit("--pairs requires --view-type materialized or both")

    xlsx_path = default_path(args.xlsx, "denorm_full_results_with_fk.xlsx")
    queries_path = default_path(args.queries_file, "workload_queries_full.sql")

    conn = R.connect_or_create(args)
    conn.autocommit = True
    cur = conn.cursor()

    if args.work_mem:
        cur.execute(f"SET work_mem = '{args.work_mem}'")
        print(f"[cfg] work_mem = {args.work_mem}")
    if args.max_parallel is not None:
        cur.execute(f"SET max_parallel_workers_per_gather = {args.max_parallel}")
        print(f"[cfg] max_parallel_workers_per_gather = {args.max_parallel}")

    nodes = R.all_nodes()
    comps_of = {i: R.components(n) for i, n in enumerate(nodes)}
    R.save_queries(str(queries_path), nodes, comps_of)

    print(f"[ok] wrote {queries_path} (Q1..Q15)")
    print(f"[cfg] view types = {modes}")
    print(f"[cfg] single-view DBs = [0] + {single_dbs}")
    print(f"[cfg] materialized pairs = {[f'{a}+{b}' for _n, (a, b) in pairs]}")
    print(f"[cfg] workload queries = {queries}")
    print(f"[cfg] pair strategy = {args.optimize}")
    print(f"[cfg] mv indexes = {args.mv_indexes}")

    records = []
    try:
        for sf in sfs:
            print(f"\n############### scale factor sf = {sf} ###############")
            c = R.build_database(
                conn, sf, args.teams, args.games_per_season,
                args.players_per_season, args.turnover, args.per_game, args.seed
            )
            print(
                f"[build] teams={c['teams']} players={c['players']} "
                f"games={c['games']} stats={c['stats']}"
            )
            base_dims = base_attr_values(c)
            base_total = sum(base_dims.values())
            print(f"[db0]  base attribute values = {base_total:,}")

            prepare_base_tables(cur)
            print(
                f"[db0]  created {len(BASE_FK_INDEXES)} missing FK indexes; "
                "VACUUM ANALYZE completed for all base tables"
            )
            R.drop_all_views(cur)
            R._base_comp_cost.clear()

            q0 = single_workload_breakdown(cur, [], comps_of, queries,
                                           args.query_repeats, args.warmup)
            total0 = sum(q0[j] for j in queries)
            print(
                f"[db0]  workload(DB0) = {total0 * 1000:.1f} ms over "
                f"{len(queries)} queries (repeats={args.query_repeats}, agg={args.agg})"
            )

            for mode in modes:
                db0_rec = make_db0_record(mode, q0, queries, base_dims)
                db0_rec["sf"] = sf
                records.append(db0_rec)

                print(f"  --- mode: {mode} ---")
                for i in single_dbs:
                    rec = run_single(
                        cur, mode, i, nodes, comps_of, q0, queries, base_dims,
                        args.query_repeats, args.warmup, args.mv_indexes
                    )
                    rec["sf"] = sf
                    records.append(rec)
                    print(
                        f"    DB{i:<2} {rec['node']:<16} "
                        f"create={rec['create_s'] * 1000:8.1f}ms "
                        f"workload={rec['workload_s'] * 1000:8.1f}ms "
                        f"(x{rec['wl_ratio']:4.2f}) "
                        f"saving={rec['saving_s'] * 1000:8.1f}ms "
                        f"redund={rec['redundancy']:5.2f}"
                    )

                if mode == "materialized":
                    for nm, dbs in pairs:
                        rec = run_pair(
                            cur, nm, dbs, nodes, comps_of, q0, queries,
                            base_total, args.query_repeats, args.warmup,
                            optimize=(args.optimize == "B"),
                            mv_indexes=args.mv_indexes,
                        )
                        rec["sf"] = sf
                        records.append(rec)
                        print(
                            f"    {rec['label']:<6} {rec['node']:<20} "
                            f"views={rec.get('n_views', '?')} "
                            f"create={rec['create_s'] * 1000:8.1f}ms "
                            f"workload={rec['workload_s'] * 1000:8.1f}ms "
                            f"(x{rec['wl_ratio']:4.2f}) "
                            f"saving={rec['saving_s'] * 1000:8.1f}ms "
                            f"redund={rec['redundancy']:5.2f}"
                        )
    finally:
        cur.close()
        conn.close()

    write_xlsx(xlsx_path, records, queries, modes)
    print(f"\n[ok] wrote {xlsx_path}")
    print("[ok] sheets: summary, per_query_time, per_query_saving, single_style_summary")


if __name__ == "__main__":
    main()
