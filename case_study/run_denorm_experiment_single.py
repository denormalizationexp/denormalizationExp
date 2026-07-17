#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_denorm_experiment.py
========================
Self-contained denormalization-lattice experiment over the sports-stats DB.
It BUILDS the original database (DB0) itself, for each scale factor in a sweep,
then runs the full 16-node lattice experiment and records `sf` as a dimension.

Pipeline per scale factor sf:
    1. (re)build DB0 = normalized schema populated to `sf` seasons of data
    2. for each lattice node i (16 = subsets of {PS,SG,Gh,Ga}):
         create the view(s) of node i  -> measure create_time
         run Q1..Q15 on DBi            -> measure workload_time
    3. tag every result row with sf

Join atoms (FK edges). Game-Team has TWO FKs -> two edges:
    PS: player JOIN stats ; SG: stats JOIN game ;
    Gh: game JOIN team(home) ; Ga: game JOIN team(away)

Answering policy (relation-level view reuse):
    On DBi a query is answered by reusing the materialized views at the
    granularity of individual relations, mimicking a real deployment in which a
    materialized view supersedes its base tables:
      * if a single view subsumes the query (atoms(query) is a subset of
        atoms(i)), it is answered directly from that view by projection
        (+ DISTINCT when the view is wider than the query);
      * otherwise, every base relation the query needs that is materialized in
        SOME view MUST be taken from a view (never the base table); relations
        supplied by the same view are read together as one block, and a join
        predicate is needed only between relations coming from different
        providers (a join internal to one view is already materialized); only
        relations covered by no view are read from base tables;
      * if no view covers any needed relation, the query is recomputed entirely
        from the base tables (identical to DB0).
    No per-query cherry-picking: the policy is fixed and applies uniformly.

Outputs:
    denorm_results.xlsx   : one sheet per view_type; rows = all (sf, DBi)
    workload_queries.sql  : the 15 workload queries (for the report)

Example:
    python run_denorm_experiment.py --sf-sweep 1,5,10 \
        --dbname sportsdb --user YOUR_USERNAME --password YOUR_PASSWORD \
        --view-type both
"""

import argparse
import csv
import io
import random
import statistics
import time
from datetime import date, timedelta
from itertools import combinations

try:
    import psycopg2
    from psycopg2 import errors as pg_errors
except ImportError:
    raise SystemExit("psycopg2 is required:  pip install psycopg2-binary")


# ========================================================================== #
#  PART A — build the original database (DB0) at a given scale factor         #
# ========================================================================== #
TEAMS = 30
GAMES_PER_SEASON = 1230
PLAYERS_PER_SEASON = 450
TURNOVER = 0.25
STATS_PER_GAME = 30


def positive_per_game(value):
    """argparse type enforcing the SG-view coverage invariant."""
    parsed = int(value)
    if parsed < 1:
        raise argparse.ArgumentTypeError(
            "--per-game must be at least 1 so every game has a stats row"
        )
    return parsed


BUILD_DDL = """
DROP TABLE IF EXISTS stats  CASCADE;
DROP TABLE IF EXISTS game   CASCADE;
DROP TABLE IF EXISTS player CASCADE;
DROP TABLE IF EXISTS team   CASCADE;

CREATE TABLE team (
    name     VARCHAR(64)  PRIMARY KEY,
    coach    VARCHAR(128),
    location VARCHAR(128)
);
CREATE TABLE player (
    name   VARCHAR(64) PRIMARY KEY,
    dob    DATE,
    height INTEGER,
    weight INTEGER
);
CREATE TABLE game (
    "date"  DATE        NOT NULL,
    h_name  VARCHAR(64) NOT NULL,
    a_name  VARCHAR(64) NOT NULL,
    h_score INTEGER,
    a_score INTEGER,
    PRIMARY KEY ("date", h_name),
    FOREIGN KEY (h_name) REFERENCES team(name),
    FOREIGN KEY (a_name) REFERENCES team(name)
);
CREATE TABLE stats (
    p_name VARCHAR(64) NOT NULL,
    "date" DATE        NOT NULL,
    h_name VARCHAR(64) NOT NULL,
    points INTEGER,
    PRIMARY KEY ("date", h_name, p_name),
    FOREIGN KEY (p_name) REFERENCES player(name),
    FOREIGN KEY ("date", h_name) REFERENCES game("date", h_name)
);
"""


def gen_teams(n):
    return [(f"Team_{i:04d}", f"Coach_{i:04d}", f"City_{i % 100:02d}")
            for i in range(n)]


def season_game_counts(sf, per_season):
    full = int(sf)
    frac = sf - full
    counts = [per_season] * full
    if frac > 0:
        counts.append(round(per_season * frac))
    counts = [c for c in counts if c > 0]
    return counts or [max(1, round(per_season * sf))]


class PlayerPool:
    def __init__(self):
        self.rows = []
        self._next = 0
        self._base_dob = date(1985, 1, 1)

    def _add(self, k):
        new = []
        for _ in range(k):
            name = f"Player_{self._next:06d}"
            dob = self._base_dob + timedelta(days=random.randint(0, 365 * 20))
            self.rows.append((name, dob.isoformat(),
                              random.randint(170, 220), random.randint(70, 130)))
            new.append(name)
            self._next += 1
        return new

    def rosters(self, num_seasons, per_season, turnover):
        active = self._add(per_season)
        out = [list(active)]
        n_new = min(round(per_season * turnover), per_season)
        for _ in range(1, num_seasons):
            if n_new > 0:
                active = active[n_new:] + self._add(n_new)
            out.append(list(active))
        return out


def gen_games_and_stats(counts, rosters, team_names, per_game):
    games, stats = [], []
    nteams = len(team_names)
    cur = date(2000, 10, 1)
    lo, hi = max(1, nteams // 3), max(1, nteams // 2)
    for s, gcount in enumerate(counts):
        roster = rosters[s]
        k_stat = min(per_game, len(roster))
        remaining = gcount
        while remaining > 0:
            k = min(remaining, random.randint(lo, hi), nteams)
            homes = random.sample(team_names, k)
            d_iso = cur.isoformat()
            for h in homes:
                a = random.choice(team_names)
                while a == h:
                    a = random.choice(team_names)
                games.append((d_iso, h, a,
                              random.randint(80, 130), random.randint(80, 130)))
                for p in random.sample(roster, k_stat):
                    stats.append((p, d_iso, h, random.randint(0, 60)))
            remaining -= len(homes)
            cur += timedelta(days=1)
    return games, stats


def validate_game_stats_coverage(game_rows, stat_rows):
    """Reject data that an INNER JOIN SG view cannot preserve at game grain."""
    covered_games = {(d, h_name) for _p, d, h_name, _points in stat_rows}
    missing_game = next(
        (
            (d, h_name)
            for d, h_name, _a_name, _h_score, _a_score in game_rows
            if (d, h_name) not in covered_games
        ),
        None,
    )
    if missing_game is not None:
        raise RuntimeError(
            "generated data violates the SG-view coverage invariant: "
            f"game {missing_game!r} has no stats row"
        )


def copy_rows(cur, table, columns, rows):
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    buf.seek(0)
    cur.copy_expert(
        f'COPY {table} ({", ".join(columns)}) FROM STDIN WITH (FORMAT csv)', buf)


def build_database(conn, sf, teams, games_per_season, players_per_season,
                   turnover, per_game, seed):
    """(Re)build and populate DB0 at scale factor sf. Returns counts dict."""
    if per_game < 1:
        raise ValueError(
            "per_game must be at least 1 because SG-based views must preserve "
            "every game used by game/team-only queries"
        )
    random.seed(seed)
    counts = season_game_counts(sf, games_per_season)
    team_rows = gen_teams(max(2, teams))
    team_names = [t[0] for t in team_rows]
    pool = PlayerPool()
    rosters = pool.rosters(len(counts), max(2, players_per_season),
                           max(0.0, min(1.0, turnover)))
    player_rows = pool.rows
    game_rows, stat_rows = gen_games_and_stats(counts, rosters, team_names,
                                               per_game)
    validate_game_stats_coverage(game_rows, stat_rows)
    cur = conn.cursor()
    cur.execute(BUILD_DDL)        # DROP ... CASCADE also removes leftover views
    copy_rows(cur, "team", ["name", "coach", "location"], team_rows)
    copy_rows(cur, "player", ["name", "dob", "height", "weight"], player_rows)
    copy_rows(cur, "game",
              ['"date"', "h_name", "a_name", "h_score", "a_score"], game_rows)
    copy_rows(cur, "stats",
              ["p_name", '"date"', "h_name", "points"], stat_rows)
    cur.close()
    return dict(teams=len(team_rows), players=len(player_rows),
                games=len(game_rows), stats=len(stat_rows))


def create_database(args):
    conn = psycopg2.connect(dbname="postgres", host=args.host, port=args.port,
                            user=args.user, password=args.password)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            try:
                cur.execute(f'CREATE DATABASE "{args.dbname}"')
                print(f"[db] created database {args.dbname!r}")
            except pg_errors.DuplicateDatabase:
                print(f"[db] database {args.dbname!r} exists, reusing")
    finally:
        conn.close()


def connect_or_create(args):
    p = dict(dbname=args.dbname, host=args.host, port=args.port,
             user=args.user, password=args.password)
    try:
        return psycopg2.connect(**p)
    except psycopg2.OperationalError as e:
        if "does not exist" in str(e).lower():
            print(f"[db] database {args.dbname!r} not found -> creating it")
            create_database(args)
            return psycopg2.connect(**p)
        raise


# ========================================================================== #
#  PART B — lattice metadata                                                  #
# ========================================================================== #
TABLES = {
    "p":  ("player", ["name", "dob", "height", "weight"]),
    "s":  ("stats",  ["p_name", "date", "h_name", "points"]),
    "g":  ("game",   ["date", "h_name", "a_name", "h_score", "a_score"]),
    "th": ("team",   ["name", "coach", "location"]),
    "ta": ("team",   ["name", "coach", "location"]),
}
ALIAS_ORDER = ["p", "s", "g", "th", "ta"]
ATOMS = {
    "PS": (("p", "s"),  'p.name = s.p_name'),
    "SG": (("s", "g"),  's."date" = g."date" AND s.h_name = g.h_name'),
    "Gh": (("g", "th"), 'g.h_name = th.name'),
    "Ga": (("g", "ta"), 'g.a_name = ta.name'),
}
ATOM_ORDER = ["PS", "SG", "Gh", "Ga"]


def real(a):
    return TABLES[a][0]


def cols(a):
    return TABLES[a][1]


def node_key(node):
    return "_".join(a for a in ATOM_ORDER if a in node) or "EMPTY"


def node_label(node):
    return " ".join(a for a in ATOM_ORDER if a in node) or "(empty = DB0)"


# Fixed DBi ordering matching the lattice figure (atom priority PS>Gh>Ga>SG,
# laid out layer by layer; within a layer the non-SG combinations precede the
# SG combinations).  Index i in this list == the DB_i number in the figure.
NODE_ORDER = [
    (),                          # 0  : empty (DB0)
    ("PS",),                     # 1  : PS
    ("Gh",),                     # 2  : Gh
    ("Ga",),                     # 3  : Ga
    ("SG",),                     # 4  : SG
    ("PS", "Gh"),                # 5  : PS . Gh
    ("PS", "Ga"),                # 6  : PS . Ga
    ("Gh", "Ga"),                # 7  : Gh . Ga
    ("PS", "SG"),                # 8  : PS . SG
    ("SG", "Gh"),                # 9  : SG . Gh
    ("SG", "Ga"),                # 10 : SG . Ga
    ("PS", "Gh", "Ga"),          # 11 : PS . Gh . Ga
    ("PS", "SG", "Gh"),          # 12 : PS . SG . Gh
    ("PS", "SG", "Ga"),          # 13 : PS . SG . Ga
    ("SG", "Gh", "Ga"),          # 14 : SG . Gh . Ga
    ("PS", "SG", "Gh", "Ga"),    # 15 : PS . SG . Gh . Ga
]


def all_nodes():
    # return the 16 nodes in the fixed figure order
    nodes = [frozenset(t) for t in NODE_ORDER]
    # sanity: must be exactly the 16 distinct subsets of the 4 atoms
    assert len(nodes) == 16 and len(set(nodes)) == 16
    allsub = {frozenset(c) for r in range(len(ATOM_ORDER) + 1)
              for c in combinations(ATOM_ORDER, r)}
    assert set(nodes) == allsub, "NODE_ORDER must be the 16 atom subsets"
    return nodes


def components(node):
    node = set(node)
    if not node:
        return []
    verts = set()
    for a in node:
        verts |= set(ATOMS[a][0])
    parent = {v: v for v in verts}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for a in node:
        x, y = ATOMS[a][0]
        parent[find(x)] = find(y)
    groups = {}
    for v in verts:
        groups.setdefault(find(v), set()).add(v)
    comps = []
    for aliases in groups.values():
        atoms_in = [a for a in node if set(ATOMS[a][0]) <= aliases]
        comps.append((aliases, atoms_in))
    comps.sort(key=lambda c: min(ALIAS_ORDER.index(x) for x in c[0]))
    return comps


def component_sql(aliases, atoms_in):
    start = min(aliases, key=lambda x: ALIAS_ORDER.index(x))
    included = {start}
    from_sql = f"{real(start)} {start}"
    joins, extra, remaining = [], [], list(atoms_in)
    while remaining:
        progressed = False
        for a in list(remaining):
            x, y = ATOMS[a][0]
            cond = ATOMS[a][1]
            if x in included and y in included:
                extra.append(cond)
                remaining.remove(a)
                progressed = True
            elif x in included:
                joins.append(f"JOIN {real(y)} {y} ON {cond}")
                included.add(y)
                remaining.remove(a)
                progressed = True
            elif y in included:
                joins.append(f"JOIN {real(x)} {x} ON {cond}")
                included.add(x)
                remaining.remove(a)
                progressed = True
        if not progressed:
            break
    select = [f'{al}."{c}" AS {al}_{c}'
              for al in sorted(aliases, key=lambda x: ALIAS_ORDER.index(x))
              for c in cols(al)]
    sql = f"SELECT {', '.join(select)} FROM {from_sql} " + " ".join(joins)
    if extra:
        sql += " WHERE " + " AND ".join(extra)
    return sql


def proj_cols(aliases):
    return [f"{al}_{c}"
            for al in sorted(aliases, key=lambda x: ALIAS_ORDER.index(x))
            for c in cols(al)]


# ========================================================================== #
#  PART C — timing + workload                                                 #
# ========================================================================== #
AGG_FUNCS = {"median": statistics.median, "mean": statistics.mean}
AGG = statistics.median        # overridden by --agg in main()


def agg_time(cur, sql, repeats, warmup):
    for _ in range(warmup):
        cur.execute(sql)
        cur.fetchone()
    samples = []
    for _ in range(repeats):
        t = time.perf_counter()
        cur.execute(sql)
        cur.fetchone()
        samples.append(time.perf_counter() - t)
    return AGG(samples)            # median or mean, per --agg


def drop_all_views(cur):
    cur.execute("SELECT matviewname FROM pg_matviews WHERE matviewname LIKE 'mv\\_%'")
    for (n,) in cur.fetchall():
        cur.execute(f"DROP MATERIALIZED VIEW IF EXISTS {n} CASCADE")
    cur.execute("SELECT viewname FROM pg_views WHERE viewname LIKE 'mv\\_%'")
    for (n,) in cur.fetchall():
        cur.execute(f"DROP VIEW IF EXISTS {n} CASCADE")


_base_comp_cost = {}


def base_comp_cost(cur, aliases, atoms_in, repeats, warmup):
    key = frozenset(atoms_in)
    if key not in _base_comp_cost:
        sql = f"SELECT count(*) FROM ({component_sql(aliases, atoms_in)}) q"
        _base_comp_cost[key] = agg_time(cur, sql, repeats, warmup)
    return _base_comp_cost[key]


def _atom_verts(atoms_iter):
    """all relation-aliases touched by a set of atoms"""
    v = set()
    for a in atoms_iter:
        v |= set(ATOMS[a][0])
    return v


def _connected(atoms_set):
    """is the sub-join-graph induced by `atoms_set` connected?

    NOTE: unused by the current relation-level plan_query (which assigns each
    relation to a view or base independently); kept as a helper in case a
    connectivity check is needed again.
    """
    atoms_set = list(atoms_set)
    if len(atoms_set) <= 1:
        return True
    verts = _atom_verts(atoms_set)
    parent = {x: x for x in verts}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    for a in atoms_set:
        x, y = ATOMS[a][0]
        parent[find(x)] = find(y)
    roots = {find(x) for x in verts}
    return len(roots) == 1


def find_cover(cj_atoms, view_comps):
    """exact/subset cover by a single view (kept for DB0 and reporting)."""
    for vc in view_comps:
        _, vi_atoms, _ = vc
        if cj_atoms <= vi_atoms:
            return vc, (cj_atoms == vi_atoms)
    return None, False


# Grain order (finest relation first) used to decide whether projecting a
# subset of a view's relations needs DISTINCT: a view's row count is fixed by
# its finest relation, so dropping that relation from the projection produces
# duplicates and requires de-duplication.
GRAIN_ORDER = ["s", "g", "p", "th", "ta"]


def _finest(aliases):
    """the finest (largest-cardinality) relation-alias among `aliases`"""
    for a in GRAIN_ORDER:
        if a in aliases:
            return a
    return None


def plan_query(cj_atoms, view_comps):
    """
    Plan how to answer one query component with atom set `cj_atoms` on a DBi
    whose materialized relations are `view_comps`.

    Relation-level view-reuse policy (mimics a real deployment in which a
    materialized view supersedes its base tables):
      * a single view whose atoms subsume the query is used directly (fast path,
        kind 'view_exact' / 'view_subset');
      * otherwise, for EACH base relation the query needs, if that relation is
        materialized in SOME view it MUST be taken from a view (never the base
        table); only relations covered by no view are read from base tables.
        Relations supplied by the same view are taken together as one block;
        when several views cover a relation the narrowest one is chosen.
      * if no view covers any needed relation, the query is computed entirely
        from base relations (kind 'base', identical to DB0).

    Returns a dict describing the plan:
        kind        : 'view_exact' | 'view_subset' | 'reuse' | 'base'
        views       : list of (view_aliases, view_atoms, mv_name) used
        picks       : list of (view_comp, frozenset(provided_aliases)) for reuse
        base_aliases: relation-aliases read from base tables (reuse)
        atoms       : the query component's atoms (for predicate emission)
        full_aliases: all relation-aliases in the final join
    """
    cj_atoms = frozenset(cj_atoms)
    cj_aliases = _atom_verts(cj_atoms)

    # quick path: a single view exactly equals or subsumes the query (atoms)
    cov, exact = find_cover(cj_atoms, view_comps)
    if cov is not None:
        return dict(kind=("view_exact" if exact else "view_subset"),
                    views=[cov], base_atoms=[],
                    full_aliases=set(cov[0]))

    # relation-level reuse: assign every needed relation to a view (if covered)
    # or to the base tables (otherwise).
    picks_map = {}                       # view-index -> set(provided aliases)
    base_aliases = set()
    for al in cj_aliases:
        covering = [(k, vc) for k, vc in enumerate(view_comps) if al in vc[0]]
        if covering:
            k, _vc = min(covering, key=lambda kv: len(kv[1][0]))  # narrowest
            picks_map.setdefault(k, set()).add(al)
        else:
            base_aliases.add(al)

    if not picks_map:                    # no view covers any needed relation
        return dict(kind="base", views=[], base_atoms=sorted(cj_atoms),
                    full_aliases=cj_aliases)

    picks = [(view_comps[k], frozenset(als))
             for k, als in sorted(picks_map.items())]
    return dict(kind="reuse",
                views=[p[0] for p in picks],
                picks=picks,
                base_aliases=base_aliases,
                atoms=sorted(cj_atoms),
                full_aliases=cj_aliases)


def view_answer_cost(cur, cj_aliases, cover, exact, repeats, warmup):
    vi_aliases, _vi_atoms, mv = cover
    if exact:
        sql = f"SELECT count(*) FROM {mv}"
    else:
        proj = ", ".join(proj_cols(cj_aliases))
        # DISTINCT when the projection drops the view's finest relation, so the
        # remaining columns would otherwise repeat.
        need_distinct = _finest(vi_aliases) not in cj_aliases
        dk = "DISTINCT " if need_distinct else ""
        sql = f"SELECT count(*) FROM (SELECT {dk}{proj} FROM {mv}) q"
    return agg_time(cur, sql, repeats, warmup)


# --- column reference resolution for reuse plans -------------------------- #
# In a base table the join predicates use  alias.col  (e.g.  g.h_name).
# Inside a materialized view, that same column is exposed as  alias_col
# (e.g.  g_h_name) by component_sql's SELECT list.  When we stitch a view
# into a larger join we expose the view under a query alias and rewrite the
# predicate column references accordingly.

def _vcol(view_alias, alias, col):
    """column reference for `alias`.`col` as exposed by view `view_alias`"""
    return f'{view_alias}.{alias}_{col}'


def plan_sql(cj_aliases, plan):
    """
    Build a counting SQL for a relation-level reuse plan.  Each chosen view is
    mounted as a derived table v0, v1, ... exposing ONLY the relations it
    supplies to the query (as alias_col), with DISTINCT when the view's finest
    relation is not among those supplied (otherwise the projected columns would
    repeat).  Relations covered by no view are added as base tables.  A join
    predicate is emitted for every query atom whose two endpoints live in
    different providers; an atom fully inside one view is already materialized
    and is skipped.
    """
    picks = plan["picks"]                    # [(view_comp, frozenset(aliases))]
    base_aliases = set(plan["base_aliases"])
    atoms = plan["atoms"]

    loc = {}                                 # alias -> ('v', vtag) | ('b',)
    from_parts = []

    for k, (vc, prov) in enumerate(picks):
        valiases, _vatoms, mv = vc
        vtag = f"v{k}"
        prov_sorted = sorted(prov, key=lambda x: ALIAS_ORDER.index(x))
        cols_sql = ", ".join(f"{al}_{c}" for al in prov_sorted for c in cols(al))
        # DISTINCT iff we drop the view's finest relation from the projection
        need_distinct = _finest(valiases) not in prov
        dk = "DISTINCT " if need_distinct else ""
        from_parts.append(f"(SELECT {dk}{cols_sql} FROM {mv}) {vtag}")
        for al in prov:
            loc[al] = ("v", vtag)

    for al in sorted(base_aliases, key=lambda x: ALIAS_ORDER.index(x)):
        loc[al] = ("b",)
        from_parts.append(f"{real(al)} {al}")

    def colref(al, col):
        if loc[al][0] == "v":
            return f'{loc[al][1]}.{al}_{col}'
        return f'{al}."{col}"'

    def atom_pred(a):
        if a == "PS":
            return f'{colref("p","name")} = {colref("s","p_name")}'
        if a == "SG":
            return (f'{colref("s","date")} = {colref("g","date")} AND '
                    f'{colref("s","h_name")} = {colref("g","h_name")}')
        if a == "Gh":
            return f'{colref("g","h_name")} = {colref("th","name")}'
        if a == "Ga":
            return f'{colref("g","a_name")} = {colref("ta","name")}'
        raise ValueError(a)

    # predicates: every query atom whose endpoints are in different providers
    preds = []
    for a in atoms:
        x, y = ATOMS[a][0]
        if x in loc and y in loc:
            if loc[x] == loc[y] and loc[x][0] == "v":
                continue                     # internal to one view -> skip
            preds.append(atom_pred(a))

    from_sql = (" CROSS JOIN ".join(from_parts)
                if len(from_parts) > 1 else from_parts[0])
    proj = ", ".join(colref(al, c)
                     for al in sorted(cj_aliases, key=lambda x: ALIAS_ORDER.index(x))
                     for c in cols(al))
    # outer DISTINCT only if the assembled relations expose 'stats' but the
    # query does not need it (defensive; per-view projections already dedup).
    avail = set(loc)
    distinct = "DISTINCT " if ("s" in avail and "s" not in cj_aliases) else ""

    inner = f"SELECT {distinct}{proj} FROM {from_sql}"
    if preds:
        inner += " WHERE " + " AND ".join(preds)
    return f"SELECT count(*) FROM ({inner}) q"


def plan_answer_cost(cur, cj_aliases, plan, repeats, warmup):
    sql = plan_sql(cj_aliases, plan)
    return agg_time(cur, sql, repeats, warmup), sql


def workload_time(cur, view_comps, comps_of, n_nodes, repeats, warmup):
    """
    Total time of Q1..Q15 on a DBi under the relation-level VIEW-REUSE policy:
      * if a query component is subsumed by a single view -> answer from that
        view (projection, + DISTINCT when the view is wider);
      * else take every needed relation that is materialized in some view from
        that view (relations of one view as a single block, + DISTINCT when its
        finest relation is dropped) and join across providers / with the base
        relations covering the remaining (uncovered) relations;
      * else (no view covers any needed relation) recompute from base tables.
    For DB0 (view_comps == []) every query falls to the base path, so this
    reduces to the normalized baseline.
    """
    total = 0.0
    for j in range(1, n_nodes):
        for cj_aliases, cj_atoms_in in comps_of[j]:
            cj_atoms = frozenset(cj_atoms_in)
            plan = plan_query(cj_atoms, view_comps)
            if plan["kind"] == "view_exact":
                total += view_answer_cost(cur, cj_aliases, plan["views"][0],
                                          True, repeats, warmup)
            elif plan["kind"] == "view_subset":
                total += view_answer_cost(cur, cj_aliases, plan["views"][0],
                                          False, repeats, warmup)
            elif plan["kind"] == "reuse":
                t, _sql = plan_answer_cost(cur, cj_aliases, plan,
                                           repeats, warmup)
                total += t
            else:  # base
                total += base_comp_cost(cur, cj_aliases, cj_atoms_in,
                                        repeats, warmup)
    return total


def save_queries(path, nodes, comps_of):
    out = ["-- Workload Q1..Q15 for the denormalization-lattice experiment.",
           "-- Qi is the join query defining the i-th lattice node (i-th view).",
           "-- Disconnected nodes are written as multiple component queries.", ""]
    for i in range(1, len(nodes)):
        comps = comps_of[i]
        disc = len(comps) > 1
        tag = "  [DISCONNECTED: %d components]" % len(comps) if disc else ""
        out.append(f"-- ===== Q{i} : node {node_label(nodes[i])}{tag} =====")
        for c_idx, (aliases, atoms_in) in enumerate(comps):
            if disc:
                out.append(f"-- Q{i} component {c_idx + 1}:")
            out.append(component_sql(aliases, atoms_in) + ";")
        out.append("")
    with open(path, "w") as f:
        f.write("\n".join(out))


def write_xlsx(path, results, modes):
    """One sheet: shared columns (sf, i, node, total_attr_values,
    redundancy_vs_db0) followed by a side-by-side block of metrics per mode."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl is required for .xlsx output: "
                         "pip install openpyxl")

    shared = ["sf", "i", "node", "total_attr_values", "redundancy_vs_db0"]
    metric = ["create_time_s", "workload_time_s", "total_exec_time_s",
              "workload_ratio_vs_db0"]

    # pivot results so both modes sit on the same (sf, i) row
    by_key = {}
    for r in results:
        d = by_key.setdefault((r["sf"], r["i"]),
                              {"node": r["node"], "total_attr": r["total_attr"],
                               "redundancy": r["redundancy"], "modes": {}})
        d["modes"][r["view_type"]] = r
    keys = sorted(by_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    hdr_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    fill1 = PatternFill("solid", fgColor="CFCFCF")
    fill2 = PatternFill("solid", fgColor="E6E6E6")
    center = Alignment(horizontal="center", vertical="center")

    # ---- two-row header: row1 = group, row2 = column names ----
    for ci, name in enumerate(shared, start=1):          # shared span both rows
        ws.cell(1, ci, name)
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)
    col = len(shared) + 1
    for mode in modes:                                   # one block per mode
        ws.cell(1, col, mode)
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col + len(metric) - 1)
        for j, m in enumerate(metric):
            ws.cell(2, col + j, m)
        col += len(metric)
    ncols = len(shared) + len(modes) * len(metric)
    for ci in range(1, ncols + 1):
        for ri in (1, 2):
            c = ws.cell(ri, ci)
            c.font = hdr_font
            c.fill = fill1 if ri == 1 else fill2
            c.alignment = center

    # ---- data rows ----
    for ridx, key in enumerate(keys):
        d = by_key[key]
        sf, i = key
        row = [sf, i, d["node"], d["total_attr"], d["redundancy"]]
        for mode in modes:
            rr = d["modes"].get(mode)
            row += ([rr["create_s"], rr["workload_s"], rr["total_s"],
                     rr["wl_ratio"]] if rr else [None, None, None, None])
        for ci, v in enumerate(row, start=1):
            ws.cell(3 + ridx, ci, v).font = body_font

    # ---- number formats ----
    for ri in range(3, 3 + len(keys)):
        ws.cell(ri, 4).number_format = "#,##0"      # total_attr_values
        ws.cell(ri, 5).number_format = "0.0000"     # redundancy
        col = len(shared) + 1
        for _ in modes:
            ws.cell(ri, col).number_format = "0.000000"      # create
            ws.cell(ri, col + 1).number_format = "0.000000"  # workload
            ws.cell(ri, col + 2).number_format = "0.000000"  # total
            ws.cell(ri, col + 3).number_format = "0.0000"    # ratio
            col += len(metric)

    widths = [6, 5, 16, 18, 18] + [16, 17, 18, 21] * len(modes)
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A3"
    wb.save(path)


def run_mode(cur, view_type, nodes, comps_of, workload0, base_dims,
             repeats, warmup):
    kw = "MATERIALIZED VIEW" if view_type == "materialized" else "VIEW"
    base_total = sum(base_dims.values())     # DB0 = all 4 base tables
    results = [dict(i=0, view_type=view_type, node="(empty)",
                    total_attr=base_total, redundancy=1.0,
                    create_s=0.0, workload_s=workload0, total_s=workload0,
                    wl_ratio=1.0)]
    print(f"  --- mode: {view_type} ---")
    for i in range(1, len(nodes)):
        node = nodes[i]
        drop_all_views(cur)
        view_comps, create_s, view_attr = [], 0.0, 0
        involved = set()                     # base tables absorbed into a view
        for c_idx, (aliases, atoms_in) in enumerate(comps_of[i]):
            mv = f"mv_{node_key(node)}_{c_idx}"
            sql = component_sql(aliases, atoms_in)
            t = time.perf_counter()
            cur.execute(f"CREATE {kw} {mv} AS {sql}")
            create_s += time.perf_counter() - t
            view_comps.append((set(aliases), frozenset(atoms_in), mv))
            cur.execute(f"SELECT count(*) FROM {mv}")
            view_attr += cur.fetchone()[0] * len(proj_cols(aliases))  # rows*cols
            involved |= {TABLES[al][0] for al in aliases}
        wl = workload_time(cur, view_comps, comps_of, len(nodes),
                           repeats, warmup)
        total = create_s + wl
        # DBi = merged view(s)  +  base tables NOT absorbed into any join
        leftover_attr = sum(v for t, v in base_dims.items() if t not in involved)
        total_attr = view_attr + leftover_attr
        results.append(dict(i=i, view_type=view_type, node=node_label(node),
                            total_attr=total_attr,
                            redundancy=total_attr / base_total,
                            create_s=create_s, workload_s=wl, total_s=total,
                            wl_ratio=wl / workload0))
        print(f"    DB{i:<2} {node_label(node):<14} create={create_s*1000:8.1f}ms "
              f"workload={wl*1000:8.1f}ms (x{wl/workload0:4.2f}) "
              f"redund={total_attr/base_total:5.2f}")
    drop_all_views(cur)
    return results


# ========================================================================== #
#  Main — sweep over scale factors                                            #
# ========================================================================== #
def main():
    ap = argparse.ArgumentParser(
        description="Denormalization lattice experiment with an sf sweep.")
    # connection
    ap.add_argument("--dbname", default="sportsdb")
    ap.add_argument("--host", default="localhost")
    ap.add_argument("--port", default=5432, type=int)
    ap.add_argument("--user", default="", help="PostgreSQL username.")   # Enter your psql user
    ap.add_argument("--password", default="", help="PostgreSQL password.")  # Enter your psql password

    # scale-factor sweep + build params
    ap.add_argument("--sf-sweep", default="10,100",
                    help="comma-separated scale factors (seasons), e.g. 1,5,10")
    ap.add_argument("--teams", type=int, default=TEAMS)
    ap.add_argument("--games-per-season", type=int, default=GAMES_PER_SEASON)
    ap.add_argument("--players-per-season", type=int, default=PLAYERS_PER_SEASON)
    ap.add_argument("--turnover", type=float, default=TURNOVER)
    ap.add_argument("--per-game", type=positive_per_game, default=STATS_PER_GAME,
                    help="stats rows generated per game; must be at least 1")
    ap.add_argument("--seed", type=int, default=42)
    # experiment params
    ap.add_argument("--view-type", choices=["materialized", "plain", "both"],
                    default="materialized")
    ap.add_argument("--query-repeats", type=int, default=10)
    ap.add_argument("--agg", choices=["median", "mean"], default="mean",
                    help="aggregate the repeated query timings by median or mean")
    ap.add_argument("--warmup", type=int, default=2)
    ap.add_argument("--work-mem", default='256MB',
                    help="session work_mem, e.g. 256MB (default: leave PG setting)")
    ap.add_argument("--max-parallel", type=int, default=0,
                    help="session max_parallel_workers_per_gather; "
                         "0 disables parallel query (default: leave PG setting)")
    ap.add_argument("--xlsx", default="denorm_results.xlsx")
    ap.add_argument("--queries-figures", default=None)
    args = ap.parse_args()

    global AGG
    AGG = AGG_FUNCS[args.agg]

    sfs = [float(x) for x in args.sf_sweep.split(",") if x.strip()]
    modes = ["materialized", "plain"] if args.view_type == "both" \
        else [args.view_type]

    conn = connect_or_create(args)
    conn.autocommit = True
    cur = conn.cursor()

    if args.work_mem:
        cur.execute(f"SET work_mem = '{args.work_mem}'")
        print(f"[cfg] work_mem = {args.work_mem}")
    if args.max_parallel is not None:
        cur.execute(f"SET max_parallel_workers_per_gather = {args.max_parallel}")
        print(f"[cfg] max_parallel_workers_per_gather = {args.max_parallel}")

    nodes = all_nodes()
    comps_of = {i: components(n) for i, n in enumerate(nodes)}
    save_queries(args.queries_file, nodes, comps_of)
    print(f"[ok] wrote {args.queries_file} (Q1..Q15)")

    all_results = []
    for sf in sfs:
        print(f"\n############### scale factor sf = {sf} ###############")
        c = build_database(conn, sf, args.teams, args.games_per_season,
                           args.players_per_season, args.turnover,
                           args.per_game, args.seed)
        print(f"[build] teams={c['teams']} players={c['players']} "
              f"games={c['games']} stats={c['stats']}")
        # per-base-table attribute counts (rows * cols); DB0 total = their sum
        base_dims = {"player": 4 * c['players'], "team": 3 * c['teams'],
                     "game": 5 * c['games'], "stats": 4 * c['stats']}
        cur.execute("ANALYZE")
        drop_all_views(cur)
        _base_comp_cost.clear()        # data changed -> invalidate cached costs

        workload0 = workload_time(cur, [], comps_of, len(nodes),
                                  args.query_repeats, args.warmup)
        print(f"[db0]  workload(DB0) = {workload0*1000:.1f} ms "
              f"(repeats={args.query_repeats}, agg={args.agg})")
        for mode in modes:
            res = run_mode(cur, mode, nodes, comps_of, workload0, base_dims,
                           args.query_repeats, args.warmup)
            for r in res:
                r["sf"] = sf
            all_results += res
    conn.close()

    write_xlsx(args.xlsx, all_results, modes)
    print(f"\n[ok] wrote {args.xlsx}")


if __name__ == "__main__":
    main()
